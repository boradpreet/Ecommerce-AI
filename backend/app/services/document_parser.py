"""Extract plain text from uploaded knowledge-base documents."""
import csv
import io
import json
import sqlite3
import tempfile
from typing import Optional


def extract_text_from_bytes(file_name: str, data: bytes) -> str:
    """Parse document bytes into plain text for RAG indexing."""
    name = (file_name or "").lower()

    if name.endswith(".txt") or name.endswith(".md"):
        return _decode_text(data)

    if name.endswith(".pdf"):
        return _parse_pdf(data, file_name)

    if name.endswith(".docx"):
        return _parse_docx(data, file_name)

    if name.endswith(".csv"):
        return _parse_csv(data, file_name)

    if name.endswith(".json"):
        return _parse_json(data, file_name)

    if name.endswith((".xlsx", ".xls")):
        return _parse_excel(data, file_name)

    if name.endswith(".db") or name.endswith(".sqlite") or name.endswith(".sqlite3"):
        return _parse_sqlite(data, file_name)

    if name.endswith(".doc"):
        text = _decode_text(data)
        if text and len(text.strip()) > 20:
            return text
        return f"[Could not parse legacy .doc file: {file_name}. Please upload .docx or .txt instead.]"

    return _decode_text(data) or f"[Unsupported file type for: {file_name}]"


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _parse_pdf(data: bytes, file_name: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        pages = []
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                pages.append(text.strip())
        if pages:
            return "\n\n".join(pages)
        return f"[PDF {file_name} contained no extractable text.]"
    except Exception as exc:
        return f"[Failed to parse PDF {file_name}: {exc}]"


def _parse_docx(data: bytes, file_name: str) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
        if paragraphs:
            return "\n\n".join(paragraphs)
        return f"[DOCX {file_name} contained no extractable text.]"
    except Exception as exc:
        return f"[Failed to parse DOCX {file_name}: {exc}]"


def _parse_csv(data: bytes, file_name: str) -> str:
    try:
        text = _decode_text(data)
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return f"[CSV {file_name} is empty.]"

        headers = rows[0]
        lines = [f"Table from {file_name}", f"Columns: {', '.join(headers)}", ""]
        for row in rows[1:501]:
            pairs = []
            for h, v in zip(headers, row):
                if v and str(v).strip():
                    pairs.append(f"{h}: {v}")
            if pairs:
                lines.append(" | ".join(pairs))
        if len(rows) > 501:
            lines.append(f"... ({len(rows) - 501} more rows truncated)")
        return "\n".join(lines)
    except Exception as exc:
        return f"[Failed to parse CSV {file_name}: {exc}]"


def _parse_json(data: bytes, file_name: str) -> str:
    try:
        text = _decode_text(data)
        parsed = json.loads(text)
        return _json_to_text(parsed, file_name)
    except Exception as exc:
        return f"[Failed to parse JSON {file_name}: {exc}]"


def _json_to_text(obj, file_name: str, depth: int = 0) -> str:
    if depth > 6:
        return str(obj)[:500]

    if isinstance(obj, dict):
        lines = [f"Data from {file_name}"]
        for key, val in obj.items():
            if isinstance(val, (dict, list)):
                lines.append(f"{key}:")
                lines.append(_json_to_text(val, file_name, depth + 1))
            else:
                lines.append(f"{key}: {val}")
        return "\n".join(lines)

    if isinstance(obj, list):
        lines = [f"Records from {file_name} ({len(obj)} items)"]
        for i, item in enumerate(obj[:200]):
            if isinstance(item, dict):
                pairs = [f"{k}: {v}" for k, v in item.items() if v is not None]
                lines.append(f"Record {i + 1}: " + " | ".join(pairs))
            else:
                lines.append(f"Record {i + 1}: {item}")
        if len(obj) > 200:
            lines.append(f"... ({len(obj) - 200} more records truncated)")
        return "\n".join(lines)

    return str(obj)


def _parse_excel(data: bytes, file_name: str) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = [f"Spreadsheet: {file_name}"]
        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            lines.append(f"\n## Sheet: {sheet_name}")
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 500:
                    lines.append("... (more rows truncated)")
                    break
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(cells)
            if not rows:
                continue
            headers = rows[0]
            lines.append(f"Columns: {', '.join(str(h) for h in headers)}")
            for row in rows[1:]:
                pairs = []
                for h, v in zip(headers, row):
                    pairs.append(f"{h}: {v}")
                lines.append(" | ".join(pairs))
        wb.close()
        return "\n".join(lines) if len(lines) > 1 else f"[Excel {file_name} contained no data.]"
    except Exception as exc:
        return f"[Failed to parse Excel {file_name}: {exc}]"


def _parse_sqlite(data: bytes, file_name: str) -> str:
    try:
        with tempfile.NamedTemporaryFile(suffix=".db", delete=True) as tmp:
            tmp.write(data)
            tmp.flush()
            conn = sqlite3.connect(tmp.name)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            tables = cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()
            lines = [f"Database: {file_name}", f"Tables: {', '.join(t['name'] for t in tables)}", ""]
            for table in tables[:10]:
                tname = table["name"]
                cols = cur.execute(f'PRAGMA table_info("{tname}")').fetchall()
                col_names = [c["name"] for c in cols]
                lines.append(f"## Table: {tname}")
                lines.append(f"Columns: {', '.join(col_names)}")
                rows = cur.execute(f'SELECT * FROM "{tname}" LIMIT 100').fetchall()
                for row in rows:
                    pairs = [f"{c}: {row[c]}" for c in col_names if row[c] is not None]
                    if pairs:
                        lines.append(" | ".join(pairs))
                count = cur.execute(f'SELECT COUNT(*) as c FROM "{tname}"').fetchone()["c"]
                if count > 100:
                    lines.append(f"... ({count - 100} more rows in {tname})")
                lines.append("")
            conn.close()
        return "\n".join(lines)
    except Exception as exc:
        return f"[Failed to parse SQLite {file_name}: {exc}]"
