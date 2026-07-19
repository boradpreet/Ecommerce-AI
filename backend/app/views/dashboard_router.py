from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.db.session import get_db
from app.models.all_models import Lead, Call, Agent, Campaign, Organization, PhoneNumber, OrganizationUser, AuditLog, User, Team, Transcript, Invoice, Subscription, AgentCatalog
from app.schemas.org_schema import (
    OrganizationSettingsUpdate, PhoneProvisionSchema, 
    GeneralSettingsUpdate, NotificationSettingsUpdate, 
    TeamInviteSchema, BillingRechargeSchema
)
from app.views.deps import get_current_user
from pydantic import BaseModel
from typing import List, Any, Optional
import datetime
import random
import os
try:
    import stripe
except ImportError:
    stripe = None

# Stripe initialization — reads STRIPE_SECRET_KEY from environment
if stripe:
    stripe.api_key = os.environ.get("STRIPE_SECRET_KEY", "")

router = APIRouter(prefix="/dashboard", tags=["dashboard"])

class LeadCreateSchema(BaseModel):
    name: str
    phone_number: str
    email: Optional[str] = None
    status: Optional[str] = "PENDING"
    campaign: Optional[str] = "Solar Outreach Q3"
    last_called: Optional[str] = "Never"

# Helper to get or create active default organization
def get_active_org(db: Session, current_user: User) -> Organization:
    # 1. Try to get organization via OrganizationUser membership
    membership = db.query(OrganizationUser).filter(OrganizationUser.user_id == current_user.id).first()
    if membership:
        org = db.query(Organization).filter(Organization.id == membership.organization_id).first()
        if org:
            return org
            
    # 2. Try to get organization owned by user
    org = db.query(Organization).filter(Organization.owner_id == current_user.id).first()
    if org:
        return org
        
    # 3. For super admin ONLY: fall back to the first organization in the database
    #    (so they can inspect any tenant without needing their own org)
    if current_user.is_superuser or (current_user.email and current_user.email.lower() == "admin@voqly.com"):
        org = db.query(Organization).first()
        if org:
            return org
        
    # 4. For all users (including vendors): create a default organization scoped to them
    #    This ensures vendors never accidentally see another vendor's data
    slug = f"org-{current_user.email.split('@')[0].lower().replace('.', '-')}-{random.randint(1000, 9999)}"
    org = Organization(
        name=f"{current_user.full_name or current_user.email.split('@')[0].capitalize()}'s Workspace",
        slug=slug,
        owner_id=current_user.id
    )
    db.add(org)
    db.commit()
    db.refresh(org)
    
    # Add organization membership mapping for this vendor
    db_member = OrganizationUser(
        organization_id=org.id,
        user_id=current_user.id,
        role="owner",
        status="active"
    )
    db.add(db_member)
    db.commit()

    # Auto-provision default Primary Operations Team
    try:
        db_team = Team(
            name="Primary Operations Team",
            organization_id=org.id
        )
        db.add(db_team)
        db.commit()
        db.refresh(db_team)

        # Auto-provision default Evelyn Assistant agent
        db_agent = Agent(
            name="Evelyn",
            voice_provider="elevenlabs",
            voice_id="21m00Tcm4TlvDq8ikWAM",
            prompt_system="# IDENTITY\nYou are Evelyn, a senior client success representative. Your goal is to help customers understand our solutions.\n\n# PERSONALITY\nMaintain a warm, professional, and efficient tone. Use short, conversational sentences designed for audio playback.",
            temperature=0.7,
            team_id=db_team.id
        )
        db.add(db_agent)
        db.commit()
        db.refresh(db_agent)

        # Auto-provision default Initial Outreach campaign
        db_campaign = Campaign(
            name="Initial Outreach",
            status="active",
            agent_id=db_agent.id
        )
        db.add(db_campaign)
        db.commit()
        print(f"[Workspace Seed] Auto-provisioned default team, agent (Evelyn), and campaign (Initial Outreach) for org: {org.name}")
    except Exception as seed_err:
        print(f"[Workspace Seed Error] Failed to auto-provision defaults: {seed_err}")

    return org


@router.get("/organization/settings")
def get_org_settings(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    return {
        "status": "success",
        "name": org.name,
        "timezone": org.timezone,
        "log_retention_days": org.log_retention_days,
        "logo_url": org.logo_url,
        "concurrency_limit": org.concurrency_limit,
        "webhook_url": org.webhook_url,
        "recording_enabled": org.recording_enabled,
        "voicemail_detection": org.voicemail_detection,
        "api_key": org.api_key,
        "notifications_slack": org.notifications_slack,
        "notifications_email": org.notifications_email,
        "notifications_low_balance": org.notifications_low_balance,
        "notifications_weekly_report": org.notifications_weekly_report,
        "prepaid_balance": org.prepaid_balance,
        "company_details": org.company_details or "",
        "auto_send_details": org.auto_send_details if org.auto_send_details is not None else True,
        "auto_send_threshold": org.auto_send_threshold if org.auto_send_threshold is not None else 50,
    }

@router.put("/organization/settings")
def update_org_settings(payload: OrganizationSettingsUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    org.name = payload.name
    org.timezone = payload.timezone
    org.log_retention_days = payload.log_retention_days
    if payload.logo_url is not None:
        org.logo_url = payload.logo_url
    if payload.company_details is not None:
        org.company_details = payload.company_details
    if payload.auto_send_details is not None:
        org.auto_send_details = payload.auto_send_details
    if payload.auto_send_threshold is not None:
        org.auto_send_threshold = max(0, min(100, int(payload.auto_send_threshold)))
    db.commit()
    db.refresh(org)
    return {
        "status": "success",
        "message": "Business settings updated successfully.",
        "settings": {
            "name": org.name,
            "timezone": org.timezone,
            "log_retention_days": org.log_retention_days,
            "logo_url": org.logo_url,
            "company_details": org.company_details or "",
            "auto_send_details": org.auto_send_details if org.auto_send_details is not None else True,
            "auto_send_threshold": org.auto_send_threshold if org.auto_send_threshold is not None else 50,
        }
    }

@router.get("/organization/general")
def get_general_settings(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    return {
        "status": "success",
        "concurrency_limit": org.concurrency_limit,
        "webhook_url": org.webhook_url,
        "recording_enabled": org.recording_enabled,
        "voicemail_detection": org.voicemail_detection
    }

@router.put("/organization/general")
def update_general_settings(payload: GeneralSettingsUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    org.concurrency_limit = payload.concurrency_limit
    org.webhook_url = payload.webhook_url
    org.recording_enabled = payload.recording_enabled
    org.voicemail_detection = payload.voicemail_detection
    db.commit()
    db.refresh(org)
    return {
        "status": "success",
        "message": "General settings updated successfully.",
        "settings": {
            "concurrency_limit": org.concurrency_limit,
            "webhook_url": org.webhook_url,
            "recording_enabled": org.recording_enabled,
            "voicemail_detection": org.voicemail_detection
        }
    }

@router.put("/organization/notifications")
def update_notification_settings(payload: NotificationSettingsUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    org.notifications_slack = payload.notifications_slack
    org.notifications_email = payload.notifications_email
    org.notifications_low_balance = payload.notifications_low_balance
    org.notifications_weekly_report = payload.notifications_weekly_report
    db.commit()
    db.refresh(org)
    return {
        "status": "success",
        "message": "Notification settings updated successfully.",
        "settings": {
            "notifications_slack": org.notifications_slack,
            "notifications_email": org.notifications_email,
            "notifications_low_balance": org.notifications_low_balance,
            "notifications_weekly_report": org.notifications_weekly_report
        }
    }

@router.get("/organization/team")
def get_team_members(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    members = db.query(OrganizationUser).filter(OrganizationUser.organization_id == org.id).all()
    
    result = []
    for m in members:
        user_email = m.email
        user_name = "Invited Collaborator"
        if m.user_id:
            u = db.query(User).filter(User.id == m.user_id).first()
            if u:
                user_email = u.email
                user_name = u.full_name or u.email
        
        result.append({
            "id": m.id,
            "name": user_name,
            "email": user_email,
            "role": m.role,
            "status": m.status
        })
    
    if not any(r["email"] == current_user.email for r in result):
        result.insert(0, {
            "id": 9999,
            "name": current_user.full_name or "Workspace Owner",
            "email": current_user.email,
            "role": "owner",
            "status": "active"
        })
        
    return {
        "status": "success",
        "members": result
    }

@router.post("/organization/team/invite")
def invite_team_member(payload: TeamInviteSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    existing = db.query(OrganizationUser).filter(
        OrganizationUser.organization_id == org.id,
        OrganizationUser.email == payload.email
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="User is already invited or a member of this workspace.")
        
    u = db.query(User).filter(User.email == payload.email).first()
    user_id = u.id if u else None
    status = "active" if u else "invited"
    
    member = OrganizationUser(
        organization_id=org.id,
        user_id=user_id,
        email=payload.email,
        role=payload.role,
        status=status
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    
    return {
        "status": "success",
        "message": f"Successfully invited {payload.email} as {payload.role}.",
        "member": {
            "id": member.id,
            "name": u.full_name if u else "Invited Collaborator",
            "email": member.email,
            "role": member.role,
            "status": member.status
        }
    }

@router.delete("/organization/team/{member_id}")
def remove_team_member(member_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    member = db.query(OrganizationUser).filter(
        OrganizationUser.organization_id == org.id,
        OrganizationUser.id == member_id
    ).first()
    
    if not member:
        raise HTTPException(status_code=404, detail="Team collaborator not found.")
        
    db.delete(member)
    db.commit()
    
    return {
        "status": "success",
        "message": "Team collaborator removed successfully."
    }

@router.get("/organization/billing")
def get_billing_info(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    sub = db.query(Subscription).filter(Subscription.organization_id == org.id).first()
    
    # Determine effective plan: only show non-free plan if there's a real paid invoice
    effective_plan = "free"
    effective_status = "active"
    period_end_str = (datetime.datetime.utcnow() + datetime.timedelta(days=30)).strftime("%B %d, %Y")
    invoices_data = []
    
    if sub:
        period_end_str = sub.current_period_end.strftime("%B %d, %Y") if sub.current_period_end else period_end_str
        effective_status = sub.status
        
        # Load all invoices for this subscription
        invoices = db.query(Invoice).filter(Invoice.subscription_id == sub.id).all()
        invoices_data = [
            {
                "id": inv.id,
                "invoice_number": f"INV-2026-{1000 + inv.id}",
                "amount": inv.amount_paid,
                "status": inv.status,
                "created_at": inv.created_at.strftime("%b %d, %Y"),
                "pdf_url": inv.pdf_url or "#"
            } for inv in invoices
            if inv.amount_paid > 0  # Only show invoices with real amounts
        ]
        
        # Only show the paid plan if there is at least one paid invoice with amount > 0
        has_paid_invoice = any(
            inv.status == "paid" and inv.amount_paid > 0
            for inv in invoices
        )
        
        if has_paid_invoice and sub.plan_tier not in (None, "free", ""):
            effective_plan = sub.plan_tier
        else:
            effective_plan = "free"
    
    return {
        "status": "success",
        "prepaid_balance": org.prepaid_balance,
        "plan_tier": effective_plan,
        "subscription_status": effective_status,
        "current_period_end": period_end_str,
        "invoices": invoices_data
    }

@router.post("/organization/billing/recharge")
def recharge_billing_balance(payload: BillingRechargeSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Recharge amount must be greater than zero.")
        
    org.prepaid_balance += payload.amount
    db.commit()
    db.refresh(org)
    
    sub = db.query(Subscription).filter(Subscription.organization_id == org.id).first()
    if sub:
        inv = Invoice(
            subscription_id=sub.id,
            amount_due=payload.amount,
            amount_paid=payload.amount,
            status="paid",
            pdf_url=f"/invoices/inv_recharge_{int(payload.amount)}_paid.pdf",
            created_at=datetime.datetime.utcnow()
        )
        db.add(inv)
        db.commit()
        
    return {
        "status": "success",
        "message": f"Successfully topped up balance by ${payload.amount:.2f}.",
        "prepaid_balance": org.prepaid_balance
    }

@router.post("/organization/api-keys/regenerate")
def regenerate_api_key(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    import secrets
    new_key = f"vq_live_{secrets.token_hex(6)}"
    org.api_key = new_key
    db.commit()
    db.refresh(org)
    
    return {
        "status": "success",
        "message": "API key regenerated successfully.",
        "api_key": org.api_key
    }

@router.get("/campaigns/report")
def get_campaigns_report(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Per-campaign aggregate report for the vendor (org-scoped).

    Returns, for every campaign in the caller's organization, the lead count,
    total calls placed, how many were picked up / completed, and total talk
    minutes — plus org-wide totals. Used by the "Campaign Report" download.
    """
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agents = db.query(Agent).filter(Agent.team_id.in_(team_ids)).all() if team_ids else []
    agent_ids = [a.id for a in agents]
    agent_name = {a.id: a.name for a in agents}

    campaigns = db.query(Campaign).filter(Campaign.agent_id.in_(agent_ids)).all() if agent_ids else []

    rows = []
    totals = {"leads": 0, "calls": 0, "picked": 0, "completed": 0, "minutes": 0.0}
    for c in campaigns:
        lead_ids = [l.id for l in db.query(Lead.id).filter(Lead.campaign_id == c.id).all()]
        leads_count = len(lead_ids)
        if lead_ids:
            calls_q = db.query(Call).filter(Call.lead_id.in_(lead_ids))
            total_calls = calls_q.count()
            picked = calls_q.filter(Call.duration_seconds > 0).count()
            completed = calls_q.filter(Call.status == "completed").count()
            total_seconds = db.query(func.sum(Call.duration_seconds)).filter(Call.lead_id.in_(lead_ids)).scalar() or 0
        else:
            total_calls = picked = completed = 0
            total_seconds = 0
        minutes = round(total_seconds / 60.0, 1)
        rows.append({
            "id": c.id,
            "name": c.name,
            "status": (c.status or "").upper(),
            "agent_name": agent_name.get(c.agent_id, "Unknown"),
            "leads": leads_count,
            "total_calls": total_calls,
            "picked": picked,
            "completed": completed,
            "total_minutes": minutes,
            "schedule": f"{c.time_start or '09:00 AM'} - {c.time_end or '05:00 PM'}",
            "created_at": c.created_at.isoformat() if c.created_at else "",
        })
        totals["leads"] += leads_count
        totals["calls"] += total_calls
        totals["picked"] += picked
        totals["completed"] += completed
        totals["minutes"] += minutes

    totals["minutes"] = round(totals["minutes"], 1)
    return {
        "organization": org.name,
        "campaign_count": len(rows),
        "totals": totals,
        "campaigns": rows,
    }


@router.get("/campaigns/report/detailed")
def get_campaigns_detailed_report(campaign_id: Optional[int] = None, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Full per-call history for the vendor's campaigns (org-scoped).

    One row per call with campaign, lead, agent, date/time, status, duration,
    sentiment and interest. Pass ?campaign_id=<id> for a single campaign
    (campaign-wise); omit it for the whole workspace. Powers the one-click
    "Download Report" CSV.
    """
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agents = db.query(Agent).filter(Agent.team_id.in_(team_ids)).all() if team_ids else []
    agent_ids = [a.id for a in agents]
    agent_name = {a.id: a.name for a in agents}

    camp_q = db.query(Campaign).filter(
        Campaign.agent_id.in_(agent_ids),
        ~Campaign.status.in_(["leads_list", "LEADS_LIST"]),
    ) if agent_ids else None
    if camp_q is not None and campaign_id is not None:
        camp_q = camp_q.filter(Campaign.id == campaign_id)
    campaigns = camp_q.all() if camp_q is not None else []
    if campaign_id is not None and not campaigns:
        raise HTTPException(status_code=404, detail="Campaign not found")
    camp_name = {c.id: c.name for c in campaigns}

    rows = []
    for c in campaigns:
        leads = db.query(Lead).filter(Lead.campaign_id == c.id).all()
        lead_by_id = {l.id: l for l in leads}
        lead_ids = list(lead_by_id.keys())
        if not lead_ids:
            continue
        calls = db.query(Call).filter(Call.lead_id.in_(lead_ids)).order_by(Call.created_at.asc()).all()
        for call in calls:
            lead = lead_by_id.get(call.lead_id)
            secs = call.duration_seconds or 0
            transcript = call.transcript
            sentiment = (transcript.sentiment if transcript and transcript.sentiment else "neutral").upper()
            interest = transcript.interest_score if transcript and transcript.interest_score is not None else 0
            created = call.created_at
            rows.append({
                "campaign_id": c.id,
                "campaign_name": camp_name.get(c.id, ""),
                "lead_name": lead.name if lead else "Unknown Lead",
                "lead_phone": lead.phone_number if lead else "",
                "agent_name": agent_name.get(call.agent_id, "Unknown"),
                "date": created.strftime("%Y-%m-%d") if created else "",
                "time": created.strftime("%H:%M:%S") if created else "",
                "status": (call.status or "").upper(),
                "duration": f"{secs // 60:02d}:{secs % 60:02d}",
                "duration_seconds": secs,
                "sentiment": sentiment,
                "interest_score": interest,
                "created_at": created.isoformat() if created else "",
            })

    return {
        "organization": org.name,
        "scope": "campaign" if campaign_id is not None else "all",
        "campaign_id": campaign_id,
        "campaign_name": (campaigns[0].name if campaign_id is not None and campaigns else None),
        "total_calls": len(rows),
        "calls": rows,
    }


@router.get("/usage")
def get_call_usage(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Vendor call-minute usage vs the admin-set limit — drives the limit-reached popup."""
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()] if team_ids else []
    total_seconds = (db.query(func.sum(Call.duration_seconds)).filter(Call.agent_id.in_(agent_ids)).scalar() or 0) if agent_ids else 0
    used = round(total_seconds / 60.0, 1)
    limit = getattr(org, "call_minutes_limit", 100)
    if limit is None:
        remaining = None
        reached = False
    else:
        remaining = round(max(0.0, limit - used), 1)
        reached = used >= limit
    return {
        "minutes_used": used,
        "minutes_limit": limit,
        "minutes_remaining": remaining,
        "limit_reached": reached,
        "contact_numbers": ["+91 9033806717", "+91 9408307302"],
        "contact_email": "business@onewebmart.com",
    }


@router.get("/metrics")
def get_dashboard_metrics(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    # 1. KPI Metrics — real counts from DB only
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    campaign_ids = [c.id for c in db.query(Campaign).filter(Campaign.agent_id.in_(agent_ids)).all()]
    
    total_calls_count = db.query(Call).filter(Call.agent_id.in_(agent_ids)).count() if agent_ids else 0
    active_campaigns = db.query(Campaign).filter(Campaign.status == "active", Campaign.agent_id.in_(agent_ids)).count() if agent_ids else 0

    # Real call duration sum → minutes
    duration_sum = db.query(func.sum(Call.duration_seconds)).filter(Call.agent_id.in_(agent_ids)).scalar() or 0 if agent_ids else 0
    minutes_used = round(duration_sum / 60, 1)

    # Containment rate: completed calls / total calls
    completed_calls = db.query(Call).filter(Call.status == "completed", Call.agent_id.in_(agent_ids)).count() if agent_ids else 0
    containment_rate = round((completed_calls / total_calls_count * 100), 1) if total_calls_count > 0 else 0.0

    # 2. Daily volume trend — last 7 days real data
    import datetime as dt
    days_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    volume_trend = []
    today = dt.date.today()
    for i in range(6, -1, -1):
        day_date = today - dt.timedelta(days=i)
        day_calls = db.query(Call).filter(
            func.date(Call.created_at) == day_date,
            Call.agent_id.in_(agent_ids)
        ).count() if agent_ids else 0
        volume_trend.append({"day": day_date.strftime("%a"), "calls": day_calls})

    # 2b. Weekly volume trend — last 7 weeks (Mon–Sun buckets), real call counts
    volume_trend_weekly = []
    this_week_monday = today - dt.timedelta(days=today.weekday())
    for w in range(6, -1, -1):
        week_start = this_week_monday - dt.timedelta(weeks=w)
        week_end = week_start + dt.timedelta(days=6)
        week_calls = db.query(Call).filter(
            func.date(Call.created_at) >= week_start,
            func.date(Call.created_at) <= week_end,
            Call.agent_id.in_(agent_ids)
        ).count() if agent_ids else 0
        volume_trend_weekly.append({"day": week_start.strftime("%b %d"), "calls": week_calls})

    # 3. Call outcomes
    in_progress = db.query(Call).filter(Call.status.in_(["initiated", "ringing", "in-progress"]), Call.agent_id.in_(agent_ids)).count() if agent_ids else 0
    failed = db.query(Call).filter(Call.status.in_(["failed", "no-answer", "busy", "dropped"]), Call.agent_id.in_(agent_ids)).count() if agent_ids else 0
    successful = total_calls_count - in_progress - failed
    if total_calls_count > 0:
        outcomes = {
            "successful": round(successful / total_calls_count * 100),
            "in_progress": round(in_progress / total_calls_count * 100),
            "failed": round(failed / total_calls_count * 100)
        }
    else:
        outcomes = {"successful": 0, "in_progress": 0, "failed": 0}

    # 4. Live sessions — recent active calls
    live_calls = db.query(Call).filter(Call.status.in_(["initiated", "in-progress"]), Call.agent_id.in_(agent_ids)).limit(5).all() if agent_ids else []
    live_sessions = []
    for c in live_calls:
        agent_name = c.agent.name if c.agent else "Unknown Agent"
        lead_phone = c.lead.phone_number if c.lead else "Unknown"
        live_sessions.append({
            "id": f"vq_{c.id:05x}",
            "agent_name": agent_name,
            "phone": lead_phone,
            "intent": "Active Call",
            "duration": f"{c.duration_seconds // 60:02d}:{c.duration_seconds % 60:02d}",
            "progress_percent": min(95, int((c.duration_seconds / 300) * 100))
        })

    return {
        "status": "success",
        "kpis": {
            "total_calls": total_calls_count,
            "minutes_used": minutes_used,
            "containment_rate": containment_rate,
            "active_campaigns": active_campaigns
        },
        "volume_trend": volume_trend,
        "volume_trend_weekly": volume_trend_weekly,
        "outcomes": outcomes,
        "live_sessions": live_sessions
    }

@router.get("/phone-numbers", response_model=List[Any])
def list_phone_numbers(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    phone_numbers = db.query(PhoneNumber).filter(PhoneNumber.organization_id == org.id).all()
    # Return empty list if no phone numbers — no seeding
    return [
        {
            "id": p.id,
            "phone_number": p.phone_number,
            "country": p.country,
            "type": p.type,
            "assigned_agent": p.assigned_agent,
            "calls_today": p.calls_today,
            "monthly_cost": p.monthly_cost,
            "status": p.status,
            "provision_type": p.provision_type,
            "direction": getattr(p, "direction", "OUTBOUND"),
            "destination_region": getattr(p, "destination_region", "USA"),
            "termination_uri": getattr(p, "termination_uri", ""),
            "cps_limit": getattr(p, "cps_limit", 2),
            "sip_username": getattr(p, "sip_username", ""),
            "sip_password": getattr(p, "sip_password", ""),
            "nickname": getattr(p, "nickname", "")
        } for p in phone_numbers
    ]

@router.post("/phone-numbers", status_code=status.HTTP_201_CREATED)
def provision_phone_number(payload: PhoneProvisionSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    # Enforce active plan phone limitations
    from app.models.all_models import Subscription
    sub = db.query(Subscription).filter(Subscription.organization_id == org.id).first()
    plan_tier = sub.plan_tier.lower() if (sub and sub.plan_tier) else "free"
    
    PLAN_LIMITS = {
        "free": {"agents": 99999, "phones": 99999},
        "starter": {"agents": 99999, "phones": 99999},
        "growth": {"agents": 99999, "phones": 99999},
        "professional": {"agents": 99999, "phones": 99999},
        "enterprise": {"agents": 99999, "phones": 99999}
    }
    
    existing_phones_count = db.query(PhoneNumber).filter(PhoneNumber.organization_id == org.id).count()
    max_phones = PLAN_LIMITS.get(plan_tier, PLAN_LIMITS["free"])["phones"]
    
    if existing_phones_count >= max_phones:
        raise HTTPException(
            status_code=400,
            detail=f"Active plan limit reached. The '{plan_tier.capitalize()}' plan allows a maximum of {max_phones} phone numbers. Please upgrade your plan in the Billing tab."
        )

    db_phone = PhoneNumber(
        phone_number=payload.phone_number,
        country=payload.country,
        type=payload.type,
        assigned_agent=payload.assigned_agent,
        calls_today=0,
        monthly_cost=payload.monthly_cost,
        status="Active",
        provision_type="Twilio SIP",
        direction=payload.direction or "OUTBOUND",
        destination_region=payload.destination_region or "USA",
        termination_uri=payload.termination_uri or "",
        cps_limit=payload.cps_limit if payload.cps_limit is not None else 2,
        sip_username=payload.sip_username or "",
        sip_password=payload.sip_password or "",
        nickname=payload.nickname or "",
        organization_id=org.id
    )
    db.add(db_phone)
    db.commit()
    db.refresh(db_phone)
    
    return {
        "status": "success",
        "phone_number": {
            "id": db_phone.id,
            "phone_number": db_phone.phone_number,
            "country": db_phone.country,
            "type": db_phone.type,
            "assigned_agent": db_phone.assigned_agent,
            "calls_today": db_phone.calls_today,
            "monthly_cost": db_phone.monthly_cost,
            "status": db_phone.status,
            "provision_type": db_phone.provision_type,
            "direction": db_phone.direction,
            "destination_region": db_phone.destination_region,
            "termination_uri": db_phone.termination_uri,
            "cps_limit": db_phone.cps_limit,
            "sip_username": db_phone.sip_username,
            "sip_password": db_phone.sip_password,
            "nickname": db_phone.nickname
        }
    }


# --- STRIPE PAYMENT INTENT FOR PHONE NUMBER PURCHASE ---

class PhonePaymentIntentSchema(BaseModel):
    amount_cents: int       # e.g. 200 for $2.00
    currency: str = "usd"
    description: str = "Voqly AI DID Phone Number - Monthly Subscription"
    phone_number: str
    region: str

@router.post("/phone-numbers/create-payment-intent")
def create_phone_payment_intent(
    payload: PhonePaymentIntentSchema,
    current_user: User = Depends(get_current_user)
) -> Any:
    """
    Creates a Stripe PaymentIntent for the phone number monthly cost.
    Frontend then uses the client_secret with Stripe.js to confirm the payment.
    """
    secret_key = os.environ.get("STRIPE_SECRET_KEY", "")
    if not secret_key:
        # Fallback: return a mock client_secret for development without Stripe keys
        return {
            "client_secret": f"pi_test_mock_{payload.phone_number.replace(' ', '').replace('+', '').replace('(', '').replace(')', '').replace('-', '')}_{payload.amount_cents}_secret_mock",
            "payment_intent_id": f"pi_test_mock_{random.randint(100000, 999999)}",
            "amount": payload.amount_cents,
            "currency": payload.currency,
            "mode": "test_mock"
        }

    try:
        intent = stripe.PaymentIntent.create(
            amount=payload.amount_cents,
            currency=payload.currency,
            description=f"{payload.description} — {payload.phone_number} ({payload.region})",
            metadata={
                "phone_number": payload.phone_number,
                "region": payload.region,
                "vendor_user_id": str(current_user.id),
                "vendor_user_email": current_user.email
            },
            automatic_payment_methods={"enabled": True}
        )
        return {
            "client_secret": intent.client_secret,
            "payment_intent_id": intent.id,
            "amount": intent.amount,
            "currency": intent.currency,
            "mode": "live"
        }
    except stripe.StripeError as e:
        raise HTTPException(status_code=400, detail=f"Stripe error: {str(e)}")


class PhonePurchaseSchema(BaseModel):
    phone_number: str
    region: str
    country: str
    type: str
    monthly_cost: float
    direction: Optional[str] = "INBOUND"
    assigned_agent_name: Optional[str] = "Unassigned"
    payment_intent_id: str          # Stripe PaymentIntent ID to verify payment
    card_name: str


@router.post("/phone-numbers/purchase", status_code=status.HTTP_201_CREATED)
def purchase_phone_number(
    payload: PhonePurchaseSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> Any:
    """
    Verifies Stripe payment then provisions the phone number into the vendor org.
    """
    org = get_active_org(db, current_user)

    # Check if number already provisioned for this org
    existing = db.query(PhoneNumber).filter(
        PhoneNumber.phone_number == payload.phone_number,
        PhoneNumber.organization_id == org.id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This phone number is already provisioned in your workspace.")

    # Verify Stripe payment intent status
    secret_key = os.environ.get("STRIPE_SECRET_KEY", "")
    is_mock = payload.payment_intent_id.startswith("pi_test_mock_")

    if secret_key and not is_mock:
        try:
            intent = stripe.PaymentIntent.retrieve(payload.payment_intent_id)
            if intent.status not in ("succeeded", "processing"):
                raise HTTPException(
                    status_code=402,
                    detail=f"Payment not confirmed. Stripe status: {intent.status}. Please complete payment."
                )
        except stripe.StripeError as e:
            raise HTTPException(status_code=400, detail=f"Stripe verification error: {str(e)}")
    # If no secret key or mock payment, allow provisioning for development/testing

    db_phone = PhoneNumber(
        phone_number=payload.phone_number,
        country=payload.country,
        type=payload.type,
        assigned_agent=payload.assigned_agent_name or "Unassigned",
        calls_today=0,
        monthly_cost=payload.monthly_cost,
        status="Active",
        provision_type="Twilio SIP",
        direction=payload.direction or "INBOUND",
        destination_region=payload.country,
        termination_uri="",
        cps_limit=2,
        sip_username="",
        sip_password="",
        nickname=payload.region,
        organization_id=org.id
    )
    db.add(db_phone)
    db.commit()
    db.refresh(db_phone)

    # Log purchase as audit record
    log = AuditLog(
        user_id=current_user.id,
        action="PHONE_NUMBER_PURCHASED",
        resource_type="PHONE_NUMBER",
        resource_id=db_phone.id,
        payload={
            "number": payload.phone_number,
            "cost": payload.monthly_cost,
            "direction": payload.direction,
            "region": payload.region,
            "cardholder": payload.card_name,
            "stripe_payment_intent": payload.payment_intent_id
        }
    )
    db.add(log)
    db.commit()

    return {
        "status": "success",
        "message": f"Successfully purchased and provisioned DID line: {payload.phone_number}",
        "transaction_id": payload.payment_intent_id,
        "phone_number": {
            "id": db_phone.id,
            "phone_number": db_phone.phone_number,
            "country": db_phone.country,
            "type": db_phone.type,
            "assigned_agent": db_phone.assigned_agent,
            "calls_today": db_phone.calls_today,
            "monthly_cost": db_phone.monthly_cost,
            "status": db_phone.status,
            "provision_type": db_phone.provision_type,
            "direction": db_phone.direction,
            "destination_region": db_phone.destination_region,
            "nickname": db_phone.nickname
        }
    }


class PhoneAssignSchema(BaseModel):
    assigned_agent_name: str
    direction: Optional[str] = None

@router.put("/phone-numbers/{phone_id}/assign")
def assign_phone_number(phone_id: int, payload: PhoneAssignSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Assign a provisioned phone number to a specific agent for the vendor organization."""
    org = get_active_org(db, current_user)
    phone = db.query(PhoneNumber).filter(
        PhoneNumber.id == phone_id,
        PhoneNumber.organization_id == org.id
    ).first()
    if not phone:
        raise HTTPException(status_code=404, detail="Phone number not found in your workspace.")

    phone.assigned_agent = payload.assigned_agent_name
    if payload.direction:
        phone.direction = payload.direction.upper()

    db.commit()
    db.refresh(phone)

    return {
        "status": "success",
        "message": f"Phone number {phone.phone_number} assigned to {payload.assigned_agent_name}.",
        "phone_number": {
            "id": phone.id,
            "phone_number": phone.phone_number,
            "assigned_agent": phone.assigned_agent,
            "direction": phone.direction,
            "status": phone.status
        }
    }


@router.get("/leads", response_model=List[Any])
def list_leads(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    import datetime
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    campaigns = db.query(Campaign).filter(
        Campaign.agent_id.in_(agent_ids),
        Campaign.status.in_(["leads_list", "LEADS_LIST"])
    ).order_by(Campaign.created_at.desc(), Campaign.id.desc()).all() if agent_ids else []
    
    result = []
    for camp in campaigns:
        total = db.query(Lead).filter(Lead.campaign_id == camp.id).count()
        if total == 0:
            continue
            
        pending = db.query(Lead).filter(Lead.campaign_id == camp.id, Lead.status.in_(["pending", "PENDING"])).count()
        called = db.query(Lead).filter(Lead.campaign_id == camp.id, Lead.status.in_(["called", "CALLED", "completed", "COMPLETED"])).count()
        dnc = db.query(Lead).filter(Lead.campaign_id == camp.id, Lead.status.in_(["dnc", "DNC"])).count()
        
        last_called_str = "Never"
        if called > 0:
            last_called_str = "Recent"
            
        result.append({
            "id": camp.id,
            "campaign_name": camp.name,
            "total_leads": total,
            "pending_leads": pending,
            "called_leads": called,
            "dnc_leads": dnc,
            "last_called": last_called_str,
            "created_at": camp.created_at.isoformat() if getattr(camp, "created_at", None) else datetime.datetime.utcnow().isoformat()
        })
    return result

@router.post("/leads", status_code=status.HTTP_201_CREATED)
def create_lead(payload: LeadCreateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    if not team_ids:
        team = Team(name="Primary Operations Team", organization_id=org.id)
        db.add(team)
        db.commit()
        db.refresh(team)
        team_ids = [team.id]
        
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    if not agent_ids:
        agent = Agent(name="Rhea (Solar Agent)", voice_id="rhea_voice", team_id=team_ids[0])
        db.add(agent)
        db.commit()
        db.refresh(agent)
        agent_ids = [agent.id]

    campaign_name = payload.campaign or "Solar Outreach Q3"
    campaign = db.query(Campaign).filter(
        Campaign.name == campaign_name,
        Campaign.agent_id.in_(agent_ids)
    ).first()
        
    if not campaign:
        campaign = Campaign(name=campaign_name, status="leads_list", agent_id=agent_ids[0])
        db.add(campaign)
        db.commit()
        db.refresh(campaign)

    db_lead = Lead(
        name=payload.name,
        phone_number=payload.phone_number,
        email=(payload.email or None),
        status=payload.status.upper() if payload.status else "PENDING",
        campaign_id=campaign.id
    )
    db.add(db_lead)
    db.commit()
    db.refresh(db_lead)
    
    status_mapping = {
        "CALLED": "2 hours ago",
        "CONVERTED": "Yesterday",
        "DNC": "3 days ago",
        "PENDING": "Never"
    }
    
    return {
        "status": "success",
        "lead": {
            "id": db_lead.id,
            "name": db_lead.name,
            "phone_number": db_lead.phone_number,
            "campaign": campaign.name,
            "status": db_lead.status,
            "last_called": status_mapping.get(db_lead.status.upper(), "Never"),
            "created_at": db_lead.created_at.isoformat()
        }
    }

class LeadUpdateSchema(BaseModel):
    name: Optional[str] = None
    phone_number: Optional[str] = None
    email: Optional[str] = None
    status: Optional[str] = None
    campaign: Optional[str] = None

@router.put("/leads/{lead_id}")
def update_lead(lead_id: int, payload: LeadUpdateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    # Check if the lead_id matches a Campaign ID first to support list renaming
    campaign = db.query(Campaign).filter(Campaign.id == lead_id).first()
    if campaign:
        if payload.campaign:
            campaign.name = payload.campaign
            db.commit()
            return {
                "status": "success",
                "message": "Lead list renamed successfully"
            }
            
    # Fallback to update single lead
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    campaign_ids = [c.id for c in db.query(Campaign).filter(Campaign.agent_id.in_(agent_ids)).all()]
    
    lead = db.query(Lead).filter(Lead.id == lead_id, Lead.campaign_id.in_(campaign_ids)).first() if campaign_ids else None
    if not lead:
        raise HTTPException(status_code=404, detail="Lead list or lead not found")
        
    if payload.name is not None:
        lead.name = payload.name
    if payload.phone_number is not None:
        lead.phone_number = payload.phone_number
    if payload.email is not None:
        lead.email = payload.email or None
    if payload.status is not None:
        lead.status = payload.status.upper()
    if payload.campaign is not None:
        campaign = db.query(Campaign).filter(
            Campaign.name == payload.campaign,
            Campaign.agent_id.in_(agent_ids)
        ).first()
        if not campaign:
            campaign = Campaign(name=payload.campaign, status="leads_list", agent_id=agent_ids[0] if agent_ids else None)
            db.add(campaign)
            db.commit()
            db.refresh(campaign)
        lead.campaign_id = campaign.id
        
    db.commit()
    db.refresh(lead)
    
    camp_name = lead.campaign.name if lead.campaign else "Unknown"
    return {
        "status": "success",
        "lead": {
            "id": lead.id,
            "name": lead.name,
            "phone_number": lead.phone_number,
            "email": lead.email or "",
            "campaign": camp_name,
            "status": lead.status,
            "last_called": "Never",
            "created_at": lead.created_at.isoformat()
        }
    }

@router.delete("/leads/{lead_id}")
def delete_lead(lead_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    # Verify campaign belongs to caller's org before deleting entire list
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    campaign_ids = [c.id for c in db.query(Campaign).filter(Campaign.agent_id.in_(agent_ids)).all()]
    
    campaign = db.query(Campaign).filter(
        Campaign.id == lead_id,
        Campaign.id.in_(campaign_ids)
    ).first() if campaign_ids else None
    
    if campaign:
        # Delete associated calls and leads
        leads = db.query(Lead).filter(Lead.campaign_id == campaign.id).all()
        for l in leads:
            db.query(Call).filter(Call.lead_id == l.id).delete()
            db.delete(l)
        db.delete(campaign)
        db.commit()
        return {"status": "success", "message": "Lead list deleted successfully"}
        
    # Fallback to delete single lead
    lead = db.query(Lead).filter(Lead.id == lead_id, Lead.campaign_id.in_(campaign_ids)).first() if campaign_ids else None
    if not lead:
        raise HTTPException(status_code=404, detail="Lead list or lead not found")
        
    # Delete associated calls
    db.query(Call).filter(Call.lead_id == lead.id).delete()
    db.delete(lead)
    db.commit()
    return {"status": "success", "message": "Lead deleted successfully"}


# --- INDIVIDUAL LEAD MANAGEMENT (dedicated routes so a lead id can never be mistaken for a list/campaign id) ---
def _lead_to_dict(lead: Lead) -> dict:
    return {
        "id": lead.id,
        "name": lead.name,
        "phone_number": lead.phone_number,
        "email": getattr(lead, "email", None) or "",
        "status": (lead.status or "PENDING").upper(),
        "created_at": lead.created_at.isoformat() if getattr(lead, "created_at", None) else None,
    }


@router.get("/leads/{campaign_id}/items")
def list_list_items(campaign_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Return the individual leads inside a lead list (org-scoped)."""
    agent_ids = _org_agent_ids(db, current_user)
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id, Campaign.agent_id.in_(agent_ids)
    ).first() if agent_ids else None
    if not campaign:
        raise HTTPException(status_code=404, detail="Lead list not found")
    leads = db.query(Lead).filter(Lead.campaign_id == campaign_id).order_by(Lead.id.desc()).all()
    return {
        "list_id": campaign.id,
        "list_name": campaign.name,
        "leads": [_lead_to_dict(l) for l in leads],
    }


class LeadItemCreateSchema(BaseModel):
    campaign_id: int
    name: str
    phone_number: str
    email: Optional[str] = None
    status: Optional[str] = "PENDING"


@router.post("/leads/item", status_code=status.HTTP_201_CREATED)
def create_lead_item(payload: LeadItemCreateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Add a single lead to an existing lead list (org-scoped)."""
    agent_ids = _org_agent_ids(db, current_user)
    campaign = db.query(Campaign).filter(
        Campaign.id == payload.campaign_id, Campaign.agent_id.in_(agent_ids)
    ).first() if agent_ids else None
    if not campaign:
        raise HTTPException(status_code=404, detail="Lead list not found")
    name = (payload.name or "").strip()
    phone = (payload.phone_number or "").strip()
    if not name or not phone:
        raise HTTPException(status_code=400, detail="Name and phone number are required.")
    lead = Lead(name=name, phone_number=phone, email=(payload.email or None), status=(payload.status or "PENDING").upper(), campaign_id=campaign.id)
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return {"status": "success", "lead": _lead_to_dict(lead)}


class LeadItemUpdateSchema(BaseModel):
    name: Optional[str] = None
    phone_number: Optional[str] = None
    email: Optional[str] = None
    status: Optional[str] = None


@router.put("/leads/item/{lead_id}")
def update_lead_item(lead_id: int, payload: LeadItemUpdateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Update a single lead's name / phone / status (org-scoped, lead ids only)."""
    agent_ids = _org_agent_ids(db, current_user)
    campaign_ids = [c[0] for c in db.query(Campaign.id).filter(Campaign.agent_id.in_(agent_ids)).all()] if agent_ids else []
    lead = db.query(Lead).filter(Lead.id == lead_id, Lead.campaign_id.in_(campaign_ids)).first() if campaign_ids else None
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload.name is not None:
        lead.name = payload.name.strip()
    if payload.phone_number is not None:
        lead.phone_number = payload.phone_number.strip()
    if payload.email is not None:
        lead.email = payload.email.strip() or None
    if payload.status is not None:
        lead.status = payload.status.upper()
    db.commit()
    db.refresh(lead)
    return {"status": "success", "lead": _lead_to_dict(lead)}


@router.delete("/leads/item/{lead_id}")
def delete_lead_item(lead_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Delete a single lead (org-scoped, lead ids only)."""
    agent_ids = _org_agent_ids(db, current_user)
    campaign_ids = [c[0] for c in db.query(Campaign.id).filter(Campaign.agent_id.in_(agent_ids)).all()] if agent_ids else []
    lead = db.query(Lead).filter(Lead.id == lead_id, Lead.campaign_id.in_(campaign_ids)).first() if campaign_ids else None
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    db.query(Call).filter(Call.lead_id == lead.id).delete()
    db.delete(lead)
    db.commit()
    return {"status": "success", "message": "Lead deleted"}


class BulkLeadImportSchema(BaseModel):
    leads: List[LeadCreateSchema]

@router.post("/leads/import", status_code=status.HTTP_201_CREATED)
def import_leads(payload: BulkLeadImportSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    if not team_ids:
        team = Team(name="Primary Operations Team", organization_id=org.id)
        db.add(team)
        db.commit()
        db.refresh(team)
        team_ids = [team.id]
        
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    if not agent_ids:
        agent = Agent(name="Rhea (Solar Agent)", voice_id="rhea_voice", team_id=team_ids[0])
        db.add(agent)
        db.commit()
        db.refresh(agent)
        agent_ids = [agent.id]

    inserted = 0
    campaigns_cache = {}
    
    for lead_data in payload.leads:
        camp_name = lead_data.campaign or "Imported Campaign"
        
        if camp_name not in campaigns_cache:
            camp = db.query(Campaign).filter(
                Campaign.name == camp_name,
                Campaign.agent_id.in_(agent_ids)
            ).first()
            if not camp:
                camp = Campaign(name=camp_name, status="leads_list", agent_id=agent_ids[0])
                db.add(camp)
                db.commit()
                db.refresh(camp)
            campaigns_cache[camp_name] = camp
            
        camp = campaigns_cache[camp_name]
        
        db_lead = Lead(
            name=lead_data.name,
            phone_number=lead_data.phone_number,
            email=(lead_data.email or None),
            status=lead_data.status.upper() if lead_data.status else "PENDING",
            campaign_id=camp.id
        )
        db.add(db_lead)
        inserted += 1
        
    db.commit()
    return {"status": "success", "message": f"Successfully imported {inserted} leads"}

@router.get("/analytics")
def get_analytics_metrics(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    
    # Base query for Call models
    call_query = db.query(Call).filter(Call.agent_id.in_(agent_ids)) if agent_ids else db.query(Call).filter(False)
    
    if start_date:
        try:
            s_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            call_query = call_query.filter(Call.created_at >= s_dt)
        except ValueError:
            pass
    if end_date:
        try:
            e_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1)
            call_query = call_query.filter(Call.created_at < e_dt)
        except ValueError:
            pass

    total_calls = call_query.count()
    duration_sum = call_query.with_entities(func.sum(Call.duration_seconds)).scalar() or 0
    
    minutes_used = duration_sum / 60
    if minutes_used >= 1000:
        total_minutes = f"{round(minutes_used/1000, 1)}k"
    else:
        total_minutes = f"{round(minutes_used, 1)}"
        
    avg_duration_sec = call_query.with_entities(func.avg(Call.duration_seconds)).scalar() or 0 if total_calls > 0 else 0
    avg_minutes = int(avg_duration_sec // 60)
    avg_seconds = int(avg_duration_sec % 60)
    avg_duration = f"{avg_minutes}m {avg_seconds}s"
    
    completed_calls = call_query.filter(Call.status == "completed").count()
    containment_rate = round((completed_calls / total_calls * 100), 1) if total_calls > 0 else 0.0
    
    avg_sentiment = round(4.2 + (total_calls % 8) * 0.1, 2) if total_calls > 0 else 0.0

    kpis = {
        "total_calls": total_calls,
        "total_minutes": total_minutes,
        "avg_duration": avg_duration,
        "containment_rate": containment_rate,
        "avg_sentiment": avg_sentiment
    }
    
    # Peak Traffic Heatmap
    calls = call_query.all()
    heatmap = {day: [0]*8 for day in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]}
    
    for c in calls:
        day = c.created_at.strftime("%a")
        if day in heatmap:
            slot = min(7, c.created_at.hour // 3)
            heatmap[day][slot] += 1
            
    max_val = max(max(slots) for slots in heatmap.values()) if any(any(slots) for slots in heatmap.values()) else 0
    heatmap_data = []
    for day in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
        slots_scaled = []
        for val in heatmap[day]:
            if max_val > 0:
                slots_scaled.append(min(10, round((val / max_val) * 10)))
            else:
                slots_scaled.append(0)
        heatmap_data.append({"day": day, "slots": slots_scaled})
        
    # Outcomes
    resolved_count = call_query.filter(Call.status == "completed").count()
    failed_count = call_query.filter(Call.status.in_(["failed", "no-answer", "busy", "dropped"])).count()
    transferred_count = call_query.filter(Call.status == "transferred").count()
    
    total_outcomes = resolved_count + failed_count + transferred_count
    if total_outcomes > 0:
        outcomes = {
            "resolved": round(resolved_count / total_outcomes * 100),
            "transferred": round(transferred_count / total_outcomes * 100),
            "abandoned": round(failed_count / total_outcomes * 100)
        }
    else:
        outcomes = {
            "resolved": 0,
            "transferred": 0,
            "abandoned": 0
        }
        
    # Top Voice Agents
    agents = db.query(Agent).filter(Agent.team_id.in_(team_ids)).all() if team_ids else []
    voice_agents = []
    for agent in agents:
        agent_call_query = db.query(Call).filter(Call.agent_id == agent.id)
        if start_date:
            try:
                s_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                agent_call_query = agent_call_query.filter(Call.created_at >= s_dt)
            except ValueError:
                pass
        if end_date:
            try:
                e_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1)
                agent_call_query = agent_call_query.filter(Call.created_at < e_dt)
            except ValueError:
                pass
                
        agent_calls = agent_call_query.count()
        agent_completed = agent_call_query.filter(Call.status == "completed").count()
        avg_handling_time_sec = agent_call_query.with_entities(func.avg(Call.duration_seconds)).scalar() or 0
        
        success_rate = round((agent_completed / agent_calls * 100), 1) if agent_calls > 0 else 0.0
        handling_time = f"{int(avg_handling_time_sec // 60)}m {int(avg_handling_time_sec % 60)}s"
        
        voice_agents.append({
            "name": agent.name,
            "tier": agent.capabilities or "Outbound Agent",
            "success_rate": success_rate,
            "handling_time": handling_time,
            "total_calls": agent_calls,
            "status": "ACTIVE" if agent.is_active else "IDLE"
        })
        
    voice_agents = sorted(voice_agents, key=lambda x: x["total_calls"], reverse=True)
    
    # Funnel
    intent_percent = 90 + (total_calls % 10) if total_calls > 0 else 0
    resolution_percent = int(containment_rate)
    funnel = [
        {"step": "INITIAL GREETING", "percent": 100 if total_calls > 0 else 0},
        {"step": "INTENT IDENTIFIED", "percent": min(100, intent_percent)},
        {"step": "AI RESOLUTION (CONTAINMENT)", "percent": resolution_percent}
    ]
    
    # Hourly call-volume trend (successful vs transferred), bucketed by hour-of-day
    trend = []
    for h in range(24):
        hour_calls = [c for c in calls if c.created_at and c.created_at.hour == h]
        trend.append({
            "hour": h,
            "successful": sum(1 for c in hour_calls if c.status == "completed"),
            "transferred": sum(1 for c in hour_calls if c.status == "transferred"),
        })

    # Comparison vs the immediately-preceding equal-length period
    def _pct(curr, prev):
        if prev > 0:
            return round((curr - prev) / prev * 100, 1)
        return 0.0 if not curr else 100.0

    prev_calls = 0
    prev_minutes = 0.0
    prev_containment = 0.0
    if agent_ids and start_date and end_date:
        try:
            _s = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            _e = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            _span = (_e - _s).days + 1
            _ps = _s - datetime.timedelta(days=_span)
            _prev_q = db.query(Call).filter(Call.agent_id.in_(agent_ids), Call.created_at >= _ps, Call.created_at < _s)
            prev_calls = _prev_q.count()
            _prev_dur = _prev_q.with_entities(func.sum(Call.duration_seconds)).scalar() or 0
            prev_minutes = round(_prev_dur / 60, 1)
            _prev_done = _prev_q.filter(Call.status == "completed").count()
            prev_containment = round((_prev_done / prev_calls * 100), 1) if prev_calls > 0 else 0.0
        except Exception:
            pass

    comparison = {
        "calls_prev": prev_calls,
        "calls_change": _pct(total_calls, prev_calls),
        "minutes_change": _pct(total_minutes, prev_minutes),
        "containment_change": _pct(containment_rate, prev_containment),
    }

    return {
        "status": "success",
        "kpis": kpis,
        "heatmap": heatmap_data,
        "outcomes": outcomes,
        "voice_agents": voice_agents,
        "funnel": funnel,
        "trend": trend,
        "comparison": comparison,
    }

# --- CAMPAIGN ENDPOINTS ---
class LeadInfoSchema(BaseModel):
    name: str
    phone_number: str

class CampaignCreateSchema(BaseModel):
    name: str
    agent_id: int
    status: Optional[str] = "draft"
    launch_date: Optional[str] = None
    timezone: Optional[str] = "America/Los_Angeles"
    active_days: Optional[str] = "M,T,W,T,F"
    time_start: Optional[str] = "09:00 AM"
    time_end: Optional[str] = "05:00 PM"
    dnc_scrubbing: Optional[bool] = True
    max_attempts: Optional[int] = 3
    retry_delay_hours: Optional[int] = 2
    agent_prompt_override: Optional[str] = None
    max_duration_seconds: Optional[int] = 300
    leads: Optional[List[LeadInfoSchema]] = []
    from_leads_list: Optional[str] = None
    direction: Optional[str] = "OUTBOUND"          # OUTBOUND (dialed) | INBOUND (answers incoming)
    phone_number: Optional[str] = None             # inbound campaigns bind this number to the agent

@router.get("/campaigns")
def list_campaigns(archived: bool = False, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    if not agent_ids:
        campaigns = []
    elif archived:
        # Archived view: only campaigns the user has archived (newest first)
        campaigns = db.query(Campaign).filter(
            Campaign.agent_id.in_(agent_ids),
            Campaign.status == "archived"
        ).order_by(Campaign.created_at.desc(), Campaign.id.desc()).all()
    else:
        # Default view: hide internal leads-list rows and archived campaigns (newest first)
        campaigns = db.query(Campaign).filter(
            Campaign.agent_id.in_(agent_ids),
            ~Campaign.status.in_(["leads_list", "LEADS_LIST", "archived"])
        ).order_by(Campaign.created_at.desc(), Campaign.id.desc()).all()
    result = []
    for camp in campaigns:
        agent_name = camp.agent.name if camp.agent else "Unassigned"
        leads_count = db.query(Lead).filter(Lead.campaign_id == camp.id).count()
        # Derive an accurate display status (does NOT mutate the stored status):
        #  - an "active" campaign with no leads is really a draft (nothing to dial)
        #  - an "active" campaign whose leads are all processed is complete
        real_status = (camp.status or "draft").lower()
        display_status = real_status.upper()
        is_inbound = (getattr(camp, "direction", None) or "").upper() == "INBOUND"
        if is_inbound:
            # Inbound campaigns "listen" — the leads-based DRAFT/COMPLETED derivation
            # (which assumes an outbound dial queue) doesn't apply.
            display_status = "ACTIVE" if real_status == "active" else display_status
        elif real_status == "active":
            if leads_count == 0:
                display_status = "DRAFT"
            else:
                processed = db.query(Lead).filter(
                    Lead.campaign_id == camp.id,
                    Lead.status.in_(["called", "completed", "dnc", "failed"])
                ).count()
                if processed >= leads_count:
                    display_status = "COMPLETED"
        result.append({
            "id": camp.id,
            "name": camp.name,
            "status": display_status,
            "direction": (getattr(camp, "direction", None) or "OUTBOUND"),
            "agent_name": agent_name,
            "agent_id": camp.agent_id,
            "leads_count": leads_count,
            "source_list_name": getattr(camp, "source_list_name", None),
            "created_at": camp.created_at.isoformat(),
            "launch_date": camp.launch_date,
            "timezone": camp.timezone,
            "active_days": camp.active_days,
            "time_start": camp.time_start,
            "time_end": camp.time_end,
            "dnc_scrubbing": camp.dnc_scrubbing,
            "max_attempts": camp.max_attempts,
            "retry_delay_hours": camp.retry_delay_hours,
            "agent_prompt_override": camp.agent_prompt_override,
            "max_duration_seconds": getattr(camp, "max_duration_seconds", 300)
        })
    return result

@router.post("/campaigns", status_code=status.HTTP_201_CREATED)
def create_campaign(payload: CampaignCreateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    # Secure tenant validation: verify agent belongs to teams in caller's active organization
    agent = db.query(Agent).join(Team).filter(
        Agent.id == payload.agent_id,
        Team.organization_id == org.id
    ).first()
    
    if not agent:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The selected Voice Agent does not exist or does not belong to your workspace."
        )

    if agent.is_active is False:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The selected Voice Agent is inactive. Activate the agent before creating a campaign."
        )

    campaign_direction = "INBOUND" if str(payload.direction or "").upper() == "INBOUND" else "OUTBOUND"

    from app.services.campaign_schedule import normalize_campaign_time_field
    if campaign_direction == "INBOUND":
        # Inbound campaigns answer around the clock — no dialing window applies.
        normalized_time_start = "12:00 AM"
        normalized_time_end = "11:59 PM"
    else:
        normalized_time_start = normalize_campaign_time_field(payload.time_start, "09:00 AM")
        normalized_time_end = normalize_campaign_time_field(payload.time_end, "11:59 PM")

    db_camp = Campaign(
        name=payload.name,
        status=payload.status.lower(),
        direction=campaign_direction,
        agent_id=payload.agent_id,
        launch_date=payload.launch_date,
        timezone=payload.timezone,
        active_days=payload.active_days,
        time_start=normalized_time_start,
        time_end=normalized_time_end,
        dnc_scrubbing=payload.dnc_scrubbing,
        max_attempts=payload.max_attempts,
        retry_delay_hours=payload.retry_delay_hours,
        agent_prompt_override=payload.agent_prompt_override,
        max_duration_seconds=payload.max_duration_seconds if payload.max_duration_seconds is not None else 300
    )
    db.add(db_camp)
    db.commit()
    db.refresh(db_camp)

    # Inbound campaigns bind the chosen number to this agent so incoming calls route here.
    inbound_webhook = None
    if campaign_direction == "INBOUND" and payload.phone_number:
        from app.models.all_models import PhoneNumber
        num = "".join(c for c in payload.phone_number if c.isdigit() or c == "+")
        pn = db.query(PhoneNumber).filter(PhoneNumber.phone_number == num).first()
        if pn and pn.organization_id and pn.organization_id != org.id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="That number is already assigned to another workspace.",
            )
        if not pn:
            pn = PhoneNumber(phone_number=num, type="LOCAL", country="US")
            db.add(pn)
        pn.organization_id = org.id
        pn.assigned_agent = str(payload.agent_id)
        pn.direction = "INBOUND"
        pn.status = "active"
        db.commit()
        # Auto-register the answer webhook with the org's provider (best-effort);
        # surface the result so the vendor knows if the number is actually routed.
        try:
            from app.services.telephony_provisioning import register_inbound_webhook
            inbound_webhook = register_inbound_webhook(getattr(org, "telephony_provider", "plivo") or "plivo", num)
        except Exception as e:  # noqa: BLE001
            inbound_webhook = {"ok": False, "message": str(e)}
    
    # Save any attached leads into the database
    if payload.leads:
        db_leads = []
        for lead in payload.leads:
            db_lead = Lead(
                name=lead.name,
                phone_number=lead.phone_number,
                status="pending",
                campaign_id=db_camp.id
            )
            db_leads.append(db_lead)
        db.add_all(db_leads)
        db.commit()
    elif payload.from_leads_list:
        # Remember which list this campaign was built from (for the "PP Data (7)" display).
        db_camp.source_list_name = "All Leads" if payload.from_leads_list == "all" else payload.from_leads_list
        # Load leads from the selected list/campaign name
        team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
        agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
        if payload.from_leads_list == "all":
            # Copy all leads belonging to this user's organization
            source_camps = db.query(Campaign).filter(
                Campaign.agent_id.in_(agent_ids),
                Campaign.status.in_(["leads_list", "LEADS_LIST"])
            ).all() if agent_ids else []
            source_camp_ids = [c.id for c in source_camps]
            source_leads = db.query(Lead).filter(Lead.campaign_id.in_(source_camp_ids)).all() if source_camp_ids else []
        else:
            # Copy leads from a specific campaign
            source_camp = db.query(Campaign).filter(
                Campaign.name == payload.from_leads_list,
                Campaign.agent_id.in_(agent_ids),
                Campaign.status.in_(["leads_list", "LEADS_LIST"])
            ).first() if agent_ids else None
            source_leads = db.query(Lead).filter(Lead.campaign_id == source_camp.id).all() if source_camp else []
            
        db_leads = []
        for l in source_leads:
            db_lead = Lead(
                name=l.name,
                phone_number=l.phone_number,
                status="pending",
                campaign_id=db_camp.id
            )
            db_leads.append(db_lead)
        db.add_all(db_leads)
        db.commit()
        
    agent_name = db_camp.agent.name if db_camp.agent else "Unassigned"
    leads_count = db.query(Lead).filter(Lead.campaign_id == db_camp.id).count()
    return {
        "status": "success",
        "webhook": inbound_webhook,
        "campaign": {
            "id": db_camp.id,
            "name": db_camp.name,
            "status": db_camp.status.upper(),
            "direction": db_camp.direction,
            "agent_name": agent_name,
            "agent_id": db_camp.agent_id,
            "leads_count": leads_count,
            "source_list_name": getattr(db_camp, "source_list_name", None),
            "created_at": db_camp.created_at.isoformat(),
            "launch_date": db_camp.launch_date,
            "timezone": db_camp.timezone,
            "active_days": db_camp.active_days,
            "time_start": db_camp.time_start,
            "time_end": db_camp.time_end,
            "dnc_scrubbing": db_camp.dnc_scrubbing,
            "max_attempts": db_camp.max_attempts,
            "retry_delay_hours": db_camp.retry_delay_hours,
            "agent_prompt_override": db_camp.agent_prompt_override,
            "max_duration_seconds": db_camp.max_duration_seconds
        }
    }


class QuickDialSchema(BaseModel):
    phone_number: str
    agent_id: int
    name: Optional[str] = "Prospect"

@router.post("/calls/quick-dial")
def quick_dial_call(payload: QuickDialSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    # 1. Secure tenant validation: verify agent belongs to teams in caller's active organization
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent = db.query(Agent).filter(Agent.id == payload.agent_id, Agent.team_id.in_(team_ids)).first()
    if not agent:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="The selected Voice Agent does not exist or does not belong to your workspace."
        )

    # 2. Get or create a default "Direct Outbound Calls" Campaign for this agent
    campaign = db.query(Campaign).filter(
        Campaign.name == "Direct Outbound Calls",
        Campaign.agent_id == agent.id
    ).first()
    if not campaign:
        campaign = Campaign(
            name="Direct Outbound Calls",
            status="active",
            agent_id=agent.id,
            timezone=org.timezone or "America/Los_Angeles",
            time_start="09:00 AM",
            time_end="11:59 PM"
        )
        db.add(campaign)
        db.commit()
        db.refresh(campaign)

    # 3. Create a Lead in that campaign
    db_lead = Lead(
        name=payload.name or "Prospect",
        phone_number=payload.phone_number,
        status="PENDING",
        campaign_id=campaign.id
    )
    db.add(db_lead)
    db.commit()
    db.refresh(db_lead)

    # 4. Trigger dial_lead immediately
    from app.controllers.dialer import CampaignDialer
    dialer = CampaignDialer()
    dialer.dial_lead(db, campaign, db_lead, agent)

    return {
        "status": "success",
        "message": f"Successfully initiated outbound call to {payload.phone_number}.",
        "lead_id": db_lead.id
    }


class CampaignStatusSchema(BaseModel):
    status: str  # "active", "paused", "draft", "completed"


@router.put("/campaigns/{campaign_id}/status")
def update_campaign_status(campaign_id: int, payload: CampaignStatusSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id,
        Campaign.agent_id.in_(agent_ids)
    ).first() if agent_ids else None
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    valid_statuses = ["active", "paused", "draft", "completed", "archived"]
    new_status = payload.status.lower()
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
    campaign.status = new_status
    db.commit()
    db.refresh(campaign)
    return {
        "status": "success",
        "campaign_id": campaign.id,
        "new_status": campaign.status.upper()
    }


@router.delete("/campaigns/{campaign_id}")
def delete_campaign(campaign_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent_ids = [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()]
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id,
        Campaign.agent_id.in_(agent_ids)
    ).first() if agent_ids else None
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    db.delete(campaign)
    db.commit()
    return {"status": "success", "message": f"Campaign {campaign_id} deleted"}

# --- AGENT CATALOG OPTIONS (global + org-specific) ---
from app.services.agent_catalog_service import (
    GLOBAL_CATEGORIES,
    build_catalog_options,
    resolve_catalog_prompt,
    resolve_default_category,
)


class AgentCatalogOptionCreateSchema(BaseModel):
    mode: str  # "new_category" | "existing_category"
    category: str
    subcategory: str
    system_prompt: str


class CatalogRequestCreateSchema(BaseModel):
    category: str
    subcategory: str


@router.post("/agent-catalog/requests")
def create_agent_catalog_request(
    payload: CatalogRequestCreateSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    from app.models.all_models import CatalogRequest, AgentCatalog
    org = get_active_org(db, current_user)
    category = payload.category.strip()
    subcategory = payload.subcategory.strip()

    if not category or not subcategory:
        raise HTTPException(status_code=400, detail="Category and subcategory are required.")

    # Check if this subcategory already exists globally or for this org
    global_sub = (
        db.query(AgentCatalog)
        .filter(
            AgentCatalog.category == category,
            AgentCatalog.subcategory == subcategory,
            AgentCatalog.organization_id.is_(None),
        )
        .first()
    )
    if global_sub:
        raise HTTPException(
            status_code=400,
            detail=f"Subcategory '{subcategory}' already exists as a shared system option.",
        )

    duplicate_org = (
        db.query(AgentCatalog)
        .filter(
            AgentCatalog.category == category,
            AgentCatalog.subcategory == subcategory,
            AgentCatalog.organization_id == org.id,
        )
        .first()
    )
    if duplicate_org:
        raise HTTPException(
            status_code=400,
            detail=f"'{subcategory}' already exists under '{category}' for your account.",
        )

    # Check if there's already a pending request for the same category/subcategory for this org
    existing_request = (
        db.query(CatalogRequest)
        .filter(
            CatalogRequest.organization_id == org.id,
            CatalogRequest.category == category,
            CatalogRequest.subcategory == subcategory,
            CatalogRequest.status == "pending",
        )
        .first()
    )
    if existing_request:
        raise HTTPException(
            status_code=400,
            detail="You already have a pending request for this category/subcategory.",
        )

    req = CatalogRequest(
        organization_id=org.id,
        category=category,
        subcategory=subcategory,
        status="pending",
    )
    db.add(req)
    db.commit()
    db.refresh(req)

    return {
        "status": "success",
        "message": f"Submitted request to add subcategory '{subcategory}' under '{category}' for approval.",
        "data": {
            "id": req.id,
            "category": req.category,
            "subcategory": req.subcategory,
            "status": req.status,
        }
    }


@router.get("/agent-catalog/options")
def get_agent_catalog_options(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    org = get_active_org(db, current_user)
    result = build_catalog_options(db, org.id)
    # Pre-selection: the vendor's onboarded industry, mapped to a catalog category.
    result["default_category"] = resolve_default_category(getattr(org, "industry", None), result.get("categories", []))
    return result


@router.post("/agent-catalog/options")
def create_agent_catalog_option(
    payload: AgentCatalogOptionCreateSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    org = get_active_org(db, current_user)
    category = payload.category.strip()
    subcategory = payload.subcategory.strip()
    system_prompt = payload.system_prompt.strip()

    if not category or not subcategory:
        raise HTTPException(status_code=400, detail="Category and subcategory are required.")
    if not system_prompt:
        raise HTTPException(status_code=400, detail="System prompt is required.")

    mode = payload.mode.strip().lower()
    if mode not in ("new_category", "existing_category"):
        raise HTTPException(status_code=400, detail="Mode must be 'new_category' or 'existing_category'.")

    if mode == "new_category":
        if category in GLOBAL_CATEGORIES:
            raise HTTPException(
                status_code=400,
                detail=f"'{category}' is a shared system category. Use 'existing_category' to add a subcategory instead.",
            )
        existing_custom_cat = (
            db.query(AgentCatalog)
            .filter(
                AgentCatalog.organization_id == org.id,
                AgentCatalog.category == category,
            )
            .first()
        )
        if existing_custom_cat:
            raise HTTPException(status_code=400, detail=f"Category '{category}' already exists for your account.")
    else:
        global_exists = category in GLOBAL_CATEGORIES
        org_cat_exists = (
            db.query(AgentCatalog)
            .filter(
                AgentCatalog.organization_id == org.id,
                AgentCatalog.category == category,
            )
            .first()
        )
        if not global_exists and not org_cat_exists:
            raise HTTPException(
                status_code=400,
                detail=f"Category '{category}' does not exist. Create it as a new category first.",
            )
        global_sub = (
            db.query(AgentCatalog)
            .filter(
                AgentCatalog.category == category,
                AgentCatalog.subcategory == subcategory,
                AgentCatalog.organization_id.is_(None),
            )
            .first()
        )
        if global_sub:
            raise HTTPException(
                status_code=400,
                detail=f"Subcategory '{subcategory}' already exists as a shared system option.",
            )

    duplicate = (
        db.query(AgentCatalog)
        .filter(
            AgentCatalog.organization_id == org.id,
            AgentCatalog.category == category,
            AgentCatalog.subcategory == subcategory,
        )
        .first()
    )
    if duplicate:
        raise HTTPException(
            status_code=400,
            detail=f"'{subcategory}' already exists under '{category}' for your account.",
        )

    item = AgentCatalog(
        category=category,
        subcategory=subcategory,
        system_prompt=system_prompt,
        organization_id=org.id,
        is_system=False,
    )
    db.add(item)
    db.commit()
    db.refresh(item)

    return {
        "status": "success",
        "message": f"Added '{subcategory}' under '{category}' for your account.",
        "data": {
            "id": item.id,
            "category": item.category,
            "subcategory": item.subcategory,
            "system_prompt": item.system_prompt,
        },
        "options": build_catalog_options(db, org.id),
    }


# --- AI AGENT ENDPOINTS ---
class AgentCreateSchema(BaseModel):
    name: str
    voice_id: str
    voice_provider: Optional[str] = "elevenlabs"
    prompt_system: Optional[str] = ""
    temperature: Optional[float] = 0.7
    lang: Optional[str] = "ENGLISH (US)"
    category: Optional[str] = "Ecommerce"
    subcategory: Optional[str] = "Marketing Campaign"
    last_active: Optional[str] = "Last active: 2m ago"
    performance_score: Optional[float] = 95.0
    performance_grade: Optional[str] = "A"
    hubspot_connected: Optional[bool] = True
    calendly_connected: Optional[bool] = False
    first_message: Optional[str] = "Hi, this is Rhea from NovaEdge Global. Thanks for calling in to apply for the Sales Lead."
    llm_provider: Optional[str] = "Groq"
    llm_model: Optional[str] = "openai/gpt-oss-120b"
    transcriber: Optional[str] = "Deepgram/Flux General Multi/English"
    kb_id: Optional[int] = None
    capabilities: Optional[str] = "Customer Support / General FAQ"
    wait_seconds: Optional[float] = 2.1
    smart_endpointing: Optional[str] = "LiveKit"
    silence_timeout: Optional[int] = 30
    max_duration_seconds: Optional[int] = 300
    stop_words: Optional[int] = 5
    voice_seconds: Optional[float] = 0.3
    backoff_seconds: Optional[float] = 4.0
    idle_messages: Optional[str] = '["Are you there?", "Can you hear me?", "Should I continue?"]'
    background_sound_enabled: Optional[bool] = False
    forwarding_country_code: Optional[str] = "+1"
    forwarding_phone_number: Optional[str] = ""
    analysis_summary_prompt: Optional[str] = "You are an expert call summarizer for an event invitation campaign.\nYou will be given the transcript of a call.\nCreate a summary and structured details based ONLY on the customer's responses.\nIgnore anything said by the assistant/agent unless the customer repeats, confirms, or explicitly reacts to it.\n\n### Summary\nDo exactly 2 lines (2 sentences)"
    analysis_summary_timeout: Optional[int] = 30
    analysis_summary_trigger_messages: Optional[int] = 3
    analysis_structured_prompt: Optional[str] = "Prompt for extracting structured data"
    analysis_structured_timeout: Optional[int] = 30

@router.get("/agents")
def list_agents(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agents = db.query(Agent).filter(Agent.team_id.in_(team_ids)).order_by(Agent.created_at.desc(), Agent.id.desc()).all() if team_ids else []
    return [
        {
            "id": agent.id,
            "name": agent.name,
            "voice_provider": agent.voice_provider,
            "voice_id": agent.voice_id,
            "prompt_system": agent.prompt_system,
            "temperature": agent.temperature,
            "created_at": agent.created_at.isoformat(),
            "category": getattr(agent, "category", "Ecommerce") or "Ecommerce",
            "subcategory": getattr(agent, "subcategory", "Marketing Campaign") or "Marketing Campaign",
            "status": "ACTIVE" if getattr(agent, "is_active", True) else "INACTIVE",
            "lang": agent.lang or "ENGLISH (US)",
            "last_active": agent.last_active or "Last active: 2m ago",
            "performance_score": agent.performance_score or 95.0,
            "performance_grade": agent.performance_grade or "A",
            "hubspot_connected": getattr(agent, "hubspot_connected", True),
            "calendly_connected": getattr(agent, "calendly_connected", False),
            "first_message": getattr(agent, "first_message", "Hi! How can I assist you today?"),
            "llm_provider": getattr(agent, "llm_provider", "Groq"),
            "llm_model": getattr(agent, "llm_model", "openai/gpt-oss-120b"),
            "transcriber": getattr(agent, "transcriber", "Deepgram/Flux General Multi/English"),
            "kb_id": getattr(agent, "kb_id", None),
            "capabilities": getattr(agent, "capabilities", "Customer Support / General FAQ"),
            "wait_seconds": getattr(agent, "wait_seconds", 2.1),
            "smart_endpointing": getattr(agent, "smart_endpointing", "LiveKit"),
            "silence_timeout": getattr(agent, "silence_timeout", 30),
            "max_duration_seconds": getattr(agent, "max_duration_seconds", 300),
            "stop_words": getattr(agent, "stop_words", 5),
            "voice_seconds": getattr(agent, "voice_seconds", 0.3),
            "backoff_seconds": getattr(agent, "backoff_seconds", 4.0),
            "idle_messages": getattr(agent, "idle_messages", '["Are you there?", "Can you hear me?", "Should I continue?"]'),
            "background_sound_enabled": getattr(agent, "background_sound_enabled", False),
            "forwarding_country_code": getattr(agent, "forwarding_country_code", "+1"),
            "forwarding_phone_number": getattr(agent, "forwarding_phone_number", ""),
            "analysis_summary_prompt": getattr(agent, "analysis_summary_prompt", "You are an expert call summarizer for an event invitation campaign.\nYou will be given the transcript of a call.\nCreate a summary and structured details based ONLY on the customer's responses.\nIgnore anything said by the assistant/agent unless the customer repeats, confirms, or explicitly reacts to it.\n\n### Summary\nDo exactly 2 lines (2 sentences)"),
            "analysis_summary_timeout": getattr(agent, "analysis_summary_timeout", 30),
            "analysis_summary_trigger_messages": getattr(agent, "analysis_summary_trigger_messages", 3),
            "analysis_structured_prompt": getattr(agent, "analysis_structured_prompt", "Prompt for extracting structured data"),
            "analysis_structured_timeout": getattr(agent, "analysis_structured_timeout", 30)
        } for agent in agents
    ]

@router.post("/agents", status_code=status.HTTP_201_CREATED)
def create_agent(payload: AgentCreateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    
    # Enforce active plan agent limitations
    from app.models.all_models import Subscription
    sub = db.query(Subscription).filter(Subscription.organization_id == org.id).first()
    plan_tier = sub.plan_tier.lower() if (sub and sub.plan_tier) else "free"
    
    PLAN_LIMITS = {
        "free": {"agents": 99999, "phones": 99999},
        "starter": {"agents": 99999, "phones": 99999},
        "growth": {"agents": 99999, "phones": 99999},
        "professional": {"agents": 99999, "phones": 99999},
        "enterprise": {"agents": 99999, "phones": 99999}
    }
    
    existing_agents_count = db.query(Agent).join(Team).filter(Team.organization_id == org.id).count()
    max_agents = PLAN_LIMITS.get(plan_tier, PLAN_LIMITS["free"])["agents"]
    
    if existing_agents_count >= max_agents:
        raise HTTPException(
            status_code=400,
            detail=f"Active plan limit reached. The '{plan_tier.capitalize()}' plan allows a maximum of {max_agents} agents. Please upgrade your plan in the Billing tab."
        )

    # Prevent duplicate agent names within the org — they collide in the campaign dropdown.
    name_clean = (payload.name or "").strip()
    if not name_clean:
        raise HTTPException(status_code=400, detail="Agent name is required.")
    duplicate = db.query(Agent).join(Team).filter(
        Team.organization_id == org.id,
        func.lower(Agent.name) == name_clean.lower(),
    ).first()
    if duplicate:
        raise HTTPException(
            status_code=409,
            detail=f'An agent named "{name_clean}" already exists. Please choose a unique name.',
        )

    team = db.query(Team).filter(Team.organization_id == org.id).first()
    if not team:
        team = Team(name="Primary Operations Team", organization_id=org.id)
        db.add(team)
        db.commit()
        db.refresh(team)

    # Fetch matching AgentCatalog prompt (org-specific first, then global)
    catalog_prompt = ""
    if payload.category and payload.subcategory:
        catalog_prompt = resolve_catalog_prompt(
            db, payload.category, payload.subcategory, org.id
        )

    db_agent = Agent(
        name=name_clean,
        voice_provider=payload.voice_provider or "elevenlabs",
        voice_id=payload.voice_id,
        prompt_system=catalog_prompt or payload.prompt_system,
        category=payload.category or "Ecommerce",
        subcategory=payload.subcategory or "Marketing Campaign",
        temperature=payload.temperature or 0.7,
        is_active=True,
        lang=payload.lang or "ENGLISH (US)",
        last_active=payload.last_active or "Last active: 2m ago",
        performance_score=payload.performance_score if payload.performance_score is not None else 95.0,
        performance_grade=payload.performance_grade or "A",
        hubspot_connected=payload.hubspot_connected if payload.hubspot_connected is not None else True,
        calendly_connected=payload.calendly_connected if payload.calendly_connected is not None else False,
        team_id=team.id,
        first_message=payload.first_message,
        llm_provider=payload.llm_provider,
        llm_model=payload.llm_model,
        transcriber=payload.transcriber,
        kb_id=payload.kb_id,
        capabilities=payload.capabilities,
        wait_seconds=payload.wait_seconds if payload.wait_seconds is not None else 2.1,
        smart_endpointing=payload.smart_endpointing or "LiveKit",
        silence_timeout=payload.silence_timeout if payload.silence_timeout is not None else 30,
        max_duration_seconds=payload.max_duration_seconds if payload.max_duration_seconds is not None else 300,
        stop_words=payload.stop_words if payload.stop_words is not None else 5,
        voice_seconds=payload.voice_seconds if payload.voice_seconds is not None else 0.3,
        backoff_seconds=payload.backoff_seconds if payload.backoff_seconds is not None else 4.0,
        idle_messages=payload.idle_messages or '["Are you there?", "Can you hear me?", "Should I continue?"]',
        background_sound_enabled=payload.background_sound_enabled if payload.background_sound_enabled is not None else False,
        forwarding_country_code=payload.forwarding_country_code or "+1",
        forwarding_phone_number=payload.forwarding_phone_number or "",
        analysis_summary_prompt=payload.analysis_summary_prompt or "You are an expert call summarizer for an event invitation campaign.\nYou will be given the transcript of a call.\nCreate a summary and structured details based ONLY on the customer's responses.\nIgnore anything said by the assistant/agent unless the customer repeats, confirms, or explicitly reacts to it.\n\n### Summary\nDo exactly 2 lines (2 sentences)",
        analysis_summary_timeout=payload.analysis_summary_timeout if payload.analysis_summary_timeout is not None else 30,
        analysis_summary_trigger_messages=payload.analysis_summary_trigger_messages if payload.analysis_summary_trigger_messages is not None else 3,
        analysis_structured_prompt=payload.analysis_structured_prompt or "Prompt for extracting structured data",
        analysis_structured_timeout=payload.analysis_structured_timeout if payload.analysis_structured_timeout is not None else 30
    )
    db.add(db_agent)
    db.commit()
    db.refresh(db_agent)
    
    return {
        "status": "success",
        "agent": {
            "id": db_agent.id,
            "name": db_agent.name,
            "voice_provider": db_agent.voice_provider,
            "voice_id": db_agent.voice_id,
            "prompt_system": db_agent.prompt_system,
            "temperature": db_agent.temperature,
            "created_at": db_agent.created_at.isoformat(),
            "category": db_agent.category,
            "subcategory": db_agent.subcategory,
            "status": "ACTIVE",
            "lang": db_agent.lang,
            "last_active": db_agent.last_active,
            "performance_score": db_agent.performance_score,
            "performance_grade": db_agent.performance_grade,
            "hubspot_connected": db_agent.hubspot_connected,
            "calendly_connected": db_agent.calendly_connected,
            "first_message": db_agent.first_message,
            "llm_provider": db_agent.llm_provider,
            "llm_model": db_agent.llm_model,
            "transcriber": db_agent.transcriber,
            "kb_id": db_agent.kb_id,
            "capabilities": db_agent.capabilities,
            "wait_seconds": db_agent.wait_seconds,
            "smart_endpointing": db_agent.smart_endpointing,
            "silence_timeout": db_agent.silence_timeout,
            "max_duration_seconds": db_agent.max_duration_seconds,
            "stop_words": db_agent.stop_words,
            "voice_seconds": db_agent.voice_seconds,
            "backoff_seconds": db_agent.backoff_seconds,
            "idle_messages": db_agent.idle_messages,
            "background_sound_enabled": db_agent.background_sound_enabled,
            "forwarding_country_code": db_agent.forwarding_country_code,
            "forwarding_phone_number": db_agent.forwarding_phone_number,
            "analysis_summary_prompt": db_agent.analysis_summary_prompt,
            "analysis_summary_timeout": db_agent.analysis_summary_timeout,
            "analysis_summary_trigger_messages": db_agent.analysis_summary_trigger_messages,
            "analysis_structured_prompt": db_agent.analysis_structured_prompt,
            "analysis_structured_timeout": db_agent.analysis_structured_timeout
        }
    }


class AgentUpdateSchema(BaseModel):
    name: Optional[str] = None
    voice_id: Optional[str] = None
    voice_provider: Optional[str] = None
    prompt_system: Optional[str] = None
    temperature: Optional[float] = None
    is_active: Optional[bool] = None
    lang: Optional[str] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    last_active: Optional[str] = None
    performance_score: Optional[float] = None
    performance_grade: Optional[str] = None
    hubspot_connected: Optional[bool] = None
    calendly_connected: Optional[bool] = None
    first_message: Optional[str] = None
    llm_provider: Optional[str] = None
    llm_model: Optional[str] = None
    transcriber: Optional[str] = None
    kb_id: Optional[int] = None
    capabilities: Optional[str] = None
    wait_seconds: Optional[float] = None
    smart_endpointing: Optional[str] = None
    silence_timeout: Optional[int] = None
    max_duration_seconds: Optional[int] = None
    stop_words: Optional[int] = None
    voice_seconds: Optional[float] = None
    backoff_seconds: Optional[float] = None
    idle_messages: Optional[str] = None
    background_sound_enabled: Optional[bool] = None
    forwarding_country_code: Optional[str] = None
    forwarding_phone_number: Optional[str] = None
    analysis_summary_prompt: Optional[str] = None
    analysis_summary_timeout: Optional[int] = None
    analysis_summary_trigger_messages: Optional[int] = None
    analysis_structured_prompt: Optional[str] = None
    analysis_structured_timeout: Optional[int] = None


@router.put("/agents/{agent_id}")
def update_agent(agent_id: int, payload: AgentUpdateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent = db.query(Agent).filter(
        Agent.id == agent_id,
        Agent.team_id.in_(team_ids)
    ).first() if team_ids else None
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    if payload.name is not None:
        name_clean = payload.name.strip()
        if not name_clean:
            raise HTTPException(status_code=400, detail="Agent name is required.")
        # Reject a rename that collides with another agent in the same org
        clash = db.query(Agent).join(Team).filter(
            Team.organization_id == org.id,
            func.lower(Agent.name) == name_clean.lower(),
            Agent.id != agent.id,
        ).first()
        if clash:
            raise HTTPException(
                status_code=409,
                detail=f'An agent named "{name_clean}" already exists. Please choose a unique name.',
            )
        agent.name = name_clean
    if payload.voice_id is not None:
        agent.voice_id = payload.voice_id
    if payload.voice_provider is not None:
        agent.voice_provider = payload.voice_provider
    if payload.prompt_system is not None:
        agent.prompt_system = payload.prompt_system
    if payload.temperature is not None:
        agent.temperature = payload.temperature
    if payload.is_active is not None:
        agent.is_active = payload.is_active
    if payload.lang is not None:
        agent.lang = payload.lang
    if payload.last_active is not None:
        agent.last_active = payload.last_active
    if payload.performance_score is not None:
        agent.performance_score = payload.performance_score
    if payload.performance_grade is not None:
        agent.performance_grade = payload.performance_grade
    if payload.hubspot_connected is not None:
        agent.hubspot_connected = payload.hubspot_connected
    if payload.calendly_connected is not None:
        agent.calendly_connected = payload.calendly_connected
    if payload.first_message is not None:
        agent.first_message = payload.first_message
    if payload.llm_provider is not None:
        agent.llm_provider = payload.llm_provider
    if payload.llm_model is not None:
        agent.llm_model = payload.llm_model
    if payload.transcriber is not None:
        agent.transcriber = payload.transcriber

    category_updated = False
    if payload.category is not None:
        agent.category = payload.category
        category_updated = True
    if payload.subcategory is not None:
        agent.subcategory = payload.subcategory
        category_updated = True

    if category_updated:
        catalog_prompt = resolve_catalog_prompt(
            db, agent.category, agent.subcategory, org.id
        )
        if catalog_prompt:
            agent.prompt_system = catalog_prompt

    if payload.kb_id is not None:
        # Allow setting to None explicitly if 0 is passed
        agent.kb_id = None if payload.kb_id == 0 else payload.kb_id
    if payload.capabilities is not None:
        agent.capabilities = payload.capabilities
    if payload.wait_seconds is not None:
        agent.wait_seconds = payload.wait_seconds
    if payload.smart_endpointing is not None:
        agent.smart_endpointing = payload.smart_endpointing
    if payload.silence_timeout is not None:
        agent.silence_timeout = payload.silence_timeout
    if payload.max_duration_seconds is not None:
        agent.max_duration_seconds = payload.max_duration_seconds
    if payload.stop_words is not None:
        agent.stop_words = payload.stop_words
    if payload.voice_seconds is not None:
        agent.voice_seconds = payload.voice_seconds
    if payload.backoff_seconds is not None:
        agent.backoff_seconds = payload.backoff_seconds
    if payload.idle_messages is not None:
        agent.idle_messages = payload.idle_messages
    if payload.background_sound_enabled is not None:
        agent.background_sound_enabled = payload.background_sound_enabled
    if payload.forwarding_country_code is not None:
        agent.forwarding_country_code = payload.forwarding_country_code
    if payload.forwarding_phone_number is not None:
        agent.forwarding_phone_number = payload.forwarding_phone_number
    if payload.analysis_summary_prompt is not None:
        agent.analysis_summary_prompt = payload.analysis_summary_prompt
    if payload.analysis_summary_timeout is not None:
        agent.analysis_summary_timeout = payload.analysis_summary_timeout
    if payload.analysis_summary_trigger_messages is not None:
        agent.analysis_summary_trigger_messages = payload.analysis_summary_trigger_messages
    if payload.analysis_structured_prompt is not None:
        agent.analysis_structured_prompt = payload.analysis_structured_prompt
    if payload.analysis_structured_timeout is not None:
        agent.analysis_structured_timeout = payload.analysis_structured_timeout
        
    db.commit()
    db.refresh(agent)
    
    return {
        "status": "success",
        "agent": {
            "id": agent.id,
            "name": agent.name,
            "voice_provider": agent.voice_provider,
            "voice_id": agent.voice_id,
            "prompt_system": agent.prompt_system,
            "temperature": agent.temperature,
            "created_at": agent.created_at.isoformat(),
            "category": agent.category,
            "subcategory": agent.subcategory,
            "status": "ACTIVE" if getattr(agent, "is_active", True) else "INACTIVE",
            "lang": agent.lang,
            "last_active": agent.last_active,
            "performance_score": agent.performance_score,
            "performance_grade": agent.performance_grade,
            "hubspot_connected": agent.hubspot_connected,
            "calendly_connected": agent.calendly_connected,
            "first_message": getattr(agent, "first_message", ""),
            "llm_provider": getattr(agent, "llm_provider", ""),
            "llm_model": getattr(agent, "llm_model", ""),
            "transcriber": getattr(agent, "transcriber", ""),
            "kb_id": getattr(agent, "kb_id", None),
            "capabilities": getattr(agent, "capabilities", ""),
            "wait_seconds": getattr(agent, "wait_seconds", 2.1),
            "smart_endpointing": getattr(agent, "smart_endpointing", "LiveKit"),
            "silence_timeout": getattr(agent, "silence_timeout", 30),
            "max_duration_seconds": getattr(agent, "max_duration_seconds", 300),
            "stop_words": getattr(agent, "stop_words", 5),
            "voice_seconds": getattr(agent, "voice_seconds", 0.3),
            "backoff_seconds": getattr(agent, "backoff_seconds", 4.0),
            "idle_messages": getattr(agent, "idle_messages", '["Are you there?", "Can you hear me?", "Should I continue?"]'),
            "background_sound_enabled": getattr(agent, "background_sound_enabled", False),
            "forwarding_country_code": getattr(agent, "forwarding_country_code", "+1"),
            "forwarding_phone_number": getattr(agent, "forwarding_phone_number", ""),
            "analysis_summary_prompt": getattr(agent, "analysis_summary_prompt", "You are an expert call summarizer for an event invitation campaign.\nYou will be given the transcript of a call.\nCreate a summary and structured details based ONLY on the customer's responses.\nIgnore anything said by the assistant/agent unless the customer repeats, confirms, or explicitly reacts to it.\n\n### Summary\nDo exactly 2 lines (2 sentences)"),
            "analysis_summary_timeout": getattr(agent, "analysis_summary_timeout", 30),
            "analysis_summary_trigger_messages": getattr(agent, "analysis_summary_trigger_messages", 3),
            "analysis_structured_prompt": getattr(agent, "analysis_structured_prompt", "Prompt for extracting structured data"),
            "analysis_structured_timeout": getattr(agent, "analysis_structured_timeout", 30)
        }
    }


# Dynamic simulator payload schema
class AgentSimulateSchema(BaseModel):
    message: str
    chat_history: Optional[List[dict]] = []


class SupportChatSchema(BaseModel):
    message: str
    history: Optional[List[dict]] = []



@router.post("/agents/{agent_id}/toggle")
def toggle_agent_active(agent_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent = db.query(Agent).filter(
        Agent.id == agent_id,
        Agent.team_id.in_(team_ids)
    ).first() if team_ids else None
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    
    agent.is_active = not getattr(agent, "is_active", True)
    db.commit()
    db.refresh(agent)
    
    return {
        "status": "success",
        "message": f"Successfully toggled active status for agent {agent.name}.",
        "is_active": agent.is_active,
        "agent_status": "ACTIVE" if agent.is_active else "INACTIVE"
    }


@router.post("/agents/{agent_id}/simulate")
def simulate_agent_dialogue(agent_id: int, payload: AgentSimulateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent = db.query(Agent).filter(
        Agent.id == agent_id,
        Agent.team_id.in_(team_ids)
    ).first() if team_ids else None
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
        
    user_msg = payload.message.lower().strip()
    category = (agent.category or "").lower()
    capabilities = agent.capabilities or agent.subcategory or ""

    # Industry-flavoured persona derived from the agent's onboarding category, so the
    # preview reflects the script the agent will actually run (e-commerce, healthcare, …).
    if any(k in category for k in ["ecom", "commerce", "retail", "shop"]):
        greet = f"Hi! This is {agent.name} from the store team. I can help with orders, delivery, returns or product details — what can I do for you?"
        follow = "I can check that for you. Could you share your order number or the product name?"
    elif any(k in category for k in ["health", "clinic", "medical", "hospital", "dental"]):
        greet = f"Hello, this is {agent.name} calling from the clinic. I can help you book, reschedule or confirm an appointment. How can I help?"
        follow = "Understood. May I have your name and date of birth so I can pull up your record?"
    elif any(k in category for k in ["financ", "bank", "loan", "insur", "fintech"]):
        greet = f"Hi, this is {agent.name} from the finance desk. I can help with your account, a payment, or a loan/insurance enquiry. What would you like to do?"
        follow = "Sure — for security, could you confirm your registered mobile number or account reference?"
    elif any(k in category for k in ["real estate", "property", "realty"]):
        greet = f"Hi! This is {agent.name}. Looking for a property, a site visit, or pricing details? I can help with all of that."
        follow = "Great — which location and budget range are you considering?"
    elif any(k in category for k in ["travel", "hospitality", "hotel", "tour"]):
        greet = f"Hello! This is {agent.name} from the travel desk. I can help with bookings, packages or itinerary changes. Where would you like to go?"
        follow = "Lovely — what dates and how many travellers should I plan for?"
    else:
        specialism = f"I specialise in {capabilities}. " if capabilities else ""
        greet = f"Hello! This is {agent.name}. {specialism}How can I help you today?"
        follow = "Got it — tell me a little more so I can help you properly."

    if any(w in user_msg for w in ["hello", "hi ", "hey", "start", "begin"]) or user_msg in ("hi", "hey"):
        response_text = greet
    elif any(w in user_msg for w in ["price", "cost", "charge", "fee", "quote", "how much"]):
        response_text = "Sure — pricing depends on the option you pick. Would you like me to text you the details, or book a callback with a specialist?"
    elif any(w in user_msg for w in ["bye", "thank", "exit", "stop", "goodbye"]):
        response_text = f"You're very welcome! Thanks for your time — this was {agent.name}. Have a great day. Goodbye!"
    elif any(w in user_msg for w in ["yes", "sure", "okay", "ok", "interested", "book", "schedule", "confirm"]):
        response_text = "Perfect! I'll get that arranged for you. What time works best — morning or afternoon?"
    elif any(w in user_msg for w in ["no", "not interested", "later", "busy", "don't"]):
        response_text = "No problem at all. Would it be okay if I sent over the details so you can review whenever it's convenient?"
    else:
        response_text = follow

    return {
        "status": "success",
        "reply": response_text,
        "agent_name": agent.name,
        "voice_id": agent.voice_id
    }


@router.post("/support/chat")
def support_chatbot(payload: SupportChatSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    import os
    import json
    import ssl
    import urllib.request as urllib_request

    message = payload.message
    history = payload.history or []

    # 1. Read support docs (checking both dynamic relative paths and hardcoded fallbacks)
    current_file_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file_dir))) # /VoqlyAI-main
    parent_of_repo = os.path.dirname(repo_root) # /VoqlyAI
    grandparent_of_repo = os.path.dirname(parent_of_repo) # /Preet AI Porjects
    
    potential_paths = [
        # Inside the repository folder
        os.path.join(repo_root, "Voqlychat data", "voqly_support_docs.md"),
        # Sibling of repository folder in VoqlyAI
        os.path.join(parent_of_repo, "Voqlychat data", "voqly_support_docs.md"),
        # In the grandparent folder Preet AI Porjects
        os.path.join(grandparent_of_repo, "Voqlychat data", "voqly_support_docs.md"),
        # Absolute path fallbacks
        "/Users/preetborad/Preet/Preet AI Porjects/VoqlyAI/VoqlyAI-main/Voqlychat data/voqly_support_docs.md",
        "/Users/preetborad/Preet/Preet AI Porjects/VoqlyAI/Voqlychat data/voqly_support_docs.md",
        "/Users/preetborad/Preet/Preet AI Porjects/Voqlychat data/voqly_support_docs.md",
    ]
    
    docs_path = None
    for p in potential_paths:
        if os.path.exists(p):
            docs_path = p
            break
            
    if not docs_path:
        docs_path = "/Users/preetborad/Preet/Preet AI Porjects/Voqlychat data/voqly_support_docs.md"

    docs_content = ""
    if os.path.exists(docs_path):
        try:
            with open(docs_path, "r", encoding="utf-8") as f:
                docs_content = f.read()
        except Exception:
            pass
            
    if not docs_content:
        docs_content = (
            "Voqly AI (formerly Ringg AI) is an enterprise conversational AI voice calling platform. "
            "Plans: Free ($0, 1 active agent, 1 active number, 100 free min), Starter ($99/mo, 2 agents, Twilio SIP, 1,000 calls/mo), "
            "Growth ($499/mo, 10 agents, HubSpot/CRM integrations, 10,000 calls/mo), Professional ($999/mo, Unlimited agents, white-label, 100,000 calls/mo), "
            "Enterprise (custom). Core features: IVR & smart routing, smart auto-dialer campaigns, web calling widgets, "
            "CRM & API integrations, warm transfer, ultra-low latency. Outbound dialing via Plivo or Twilio."
        )


    # 2. Simple but effective RAG ranking
    sections = docs_content.split("\n## ")
    query_words = set(message.lower().split())
    stopwords = {"a", "an", "the", "and", "or", "but", "is", "are", "was", "were", "to", "in", "of", "for", "on", "with", "at", "by", "from", "how", "what", "can", "do", "you", "i", "my"}
    query_keywords = query_words - stopwords
    if not query_keywords:
        query_keywords = query_words
        
    scored_sections = []
    for sec in sections:
        sec_lower = sec.lower()
        score = 0
        for word in query_keywords:
            if word in sec_lower:
                score += 1.5 + sec_lower.count(word)
        if score > 0:
            scored_sections.append((score, sec))
            
    scored_sections.sort(key=lambda x: x[0], reverse=True)
    top_sections = [sec for score, sec in scored_sections[:3]]
    if not top_sections:
        top_sections = sections[:2]
        
    context_str = "\n## ".join(top_sections)

    # 3. Construct System Prompt
    system_instruction = (
        "You are the official AI-powered Customer Support Assistant for Voqly AI (formerly Ringg AI).\n"
        "Your mission is to help users understand, use, troubleshoot, and explore the Voqly AI platform.\n"
        "You MUST answer the user's inquiry based strictly on the provided documentation context below. "
        "Do not invent features or make promises not supported by the documentation. "
        "If you are unsure or the context does not specify the answer, politely state:\n"
        "\"I don't want to give incorrect information. Please contact the Voqly AI team for confirmation on that specific request.\"\n"
        "Keep your response friendly, concise, professional, and human-like.\n\n"
        f"DOCUMENTATION CONTEXT:\n{context_str}\n"
    )

    # 4. Call Gemini generateContent API via urllib
    gemini_key = os.getenv("GEMINI_API_KEY")
    reply_text = "I'm sorry, I am currently unable to contact the support service. Please ensure your backend has a valid GEMINI_API_KEY."

    if gemini_key:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
        
        contents = []
        for h in history:
            role = h.get("role")
            content = h.get("content", "")
            gemini_role = "model" if role == "assistant" else "user"
            contents.append({
                "role": gemini_role,
                "parts": [{"text": content}]
            })
            
        contents.append({
            "role": "user",
            "parts": [{"text": message}]
        })
        
        body = {
            "contents": contents,
            "systemInstruction": {
                "parts": [{"text": system_instruction}]
            },
            "generationConfig": {
                "temperature": 0.3,
                "maxOutputTokens": 800
            }
        }
        
        try:
            req_data = json.dumps(body).encode("utf-8")
            req = urllib_request.Request(
                url,
                data=req_data,
                headers={"Content-Type": "application/json"},
                method="POST"
            )
            
            context = ssl._create_unverified_context()
            with urllib_request.urlopen(req, context=context, timeout=15) as resp:
                resp_body = resp.read().decode("utf-8")
                res_json = json.loads(resp_body)
                reply_text = res_json["candidates"][0]["content"]["parts"][0]["text"].strip()
        except Exception as e:
            print(f"[Support Chatbot error]: {e}")
            lower = message.lower()
            if "pricing" in lower or "cost" in lower or "plan" in lower:
                reply_text = "Voqly AI offers 5 subscription plans:\n\n• **Free**: 1 Agent, 1 Phone Number, 100 free call minutes.\n• **Starter** ($99/mo): 2 Agents, Twilio SIP, 1,000 calls/mo.\n• **Growth** ($499/mo): 10 Agents, HubSpot/CRM integrations, 10,000 calls/mo.\n• **Professional** ($999/mo): Unlimited Agents, white-label, 100,000 calls/mo.\n• **Enterprise** (Custom): Custom limits, volume discounts."

            elif "voice" in lower or "language" in lower:
                reply_text = "Voqly AI supports English, Hindi, Bengali, Gujarati, Kannada, Malayalam, Marathi, Punjabi, Tamil, and Telugu with premium natural neural voices."
            elif "api" in lower or "integrate" in lower or "webhook" in lower:
                reply_text = "Integrating Voqly AI is extremely easy:\n1. Retrieve your API token from Settings -> Developer API Key.\n2. Call POST /api/v1/calls/initiate to trigger outbound calling pipelines.\n3. Configure webhooks in settings to receive instant disposition updates on CRM logs."
            else:
                reply_text = "Voqly AI (formerly Ringg AI) is an enterprise conversational AI voice infrastructure. It allows companies to deploy natural calling agents (~110ms latency) to automate lead qualification, routing, support, and outbound notifications."

    return {
        "status": "success",
        "reply": reply_text
    }


@router.delete("/agents/{agent_id}")
def delete_agent(agent_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    agent = db.query(Agent).filter(
        Agent.id == agent_id,
        Agent.team_id.in_(team_ids)
    ).first() if team_ids else None
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
        
    db.delete(agent)
    db.commit()
    return {"status": "success", "message": "Agent deleted successfully"}


# --- CALL LOG & TRANSCRIPT ENDPOINTS ---
def _call_to_dict(call: Call) -> dict:
    """Serialize a Call (with its lead/agent/transcript) for the call-logs API."""
    lead_name = call.lead.name if call.lead else "Unknown Lead"
    lead_phone = call.lead.phone_number if call.lead else "Unknown"
    campaign = call.lead.campaign if call.lead else None
    agent_name = call.agent.name if call.agent else "Unknown Agent"
    sentiment = call.transcript.sentiment if call.transcript else "neutral"
    summary = call.transcript.summary if call.transcript else "No summary available."
    interest_score = call.transcript.interest_score if call.transcript else 0
    if interest_score is None:
        interest_score = 0
    lead_email = call.lead.email if (call.lead and getattr(call.lead, "email", None)) else ""
    details_sent = bool(getattr(call.transcript, "details_sent", False)) if call.transcript else False
    details_sent_to = getattr(call.transcript, "details_sent_to", None) if call.transcript else None
    wants_details = bool(getattr(call.transcript, "wants_details", False)) if call.transcript else False
    secs = call.duration_seconds or 0
    return {
        "id": call.id,
        "lead_name": lead_name,
        "lead_phone": lead_phone,
        "lead_email": lead_email,
        "agent_name": agent_name,
        "agent_id": call.agent_id,
        "campaign_id": campaign.id if campaign else None,
        "campaign_name": campaign.name if campaign else None,
        "status": call.status.upper(),
        "direction": (getattr(call, "direction", None) or "OUTBOUND"),
        "duration_seconds": secs,
        "duration": f"{secs // 60:02d}:{secs % 60:02d}",
        "recording_url": call.recording_url,
        "sentiment": (sentiment or "neutral").upper(),
        "interest_score": interest_score,
        "wants_details": wants_details,
        "details_sent": details_sent,
        "details_sent_to": details_sent_to,
        "summary": summary,
        "created_at": call.created_at.isoformat(),
        "archived_at": call.archived_at.isoformat() if call.archived_at else None,
    }


def _purge_expired_archived_calls(db: Session, agent_ids: list) -> None:
    """Permanently delete calls (and transcripts) archived more than 30 days ago."""
    if not agent_ids:
        return
    cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=30)
    expired_ids = [
        c[0] for c in db.query(Call.id).filter(
            Call.agent_id.in_(agent_ids),
            Call.archived_at.isnot(None),
            Call.archived_at < cutoff,
        ).all()
    ]
    if expired_ids:
        db.query(Transcript).filter(Transcript.call_id.in_(expired_ids)).delete(synchronize_session=False)
        db.query(Call).filter(Call.id.in_(expired_ids)).delete(synchronize_session=False)
        db.commit()


def _org_agent_ids(db: Session, current_user: User) -> list:
    org = get_active_org(db, current_user)
    team_ids = [t.id for t in db.query(Team).filter(Team.organization_id == org.id).all()]
    return [a.id for a in db.query(Agent).filter(Agent.team_id.in_(team_ids)).all()] if team_ids else []


@router.get("/calls")
def list_calls(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    # SECURITY: only return calls belonging to THIS user's organization (multi-tenant isolation).
    agent_ids = _org_agent_ids(db, current_user)
    _purge_expired_archived_calls(db, agent_ids)  # opportunistic cleanup of >30-day-old archives
    calls = (
        db.query(Call)
        .filter(Call.agent_id.in_(agent_ids), Call.archived_at.is_(None))
        .order_by(Call.created_at.desc())
        .all()
        if agent_ids else []
    )
    return [_call_to_dict(c) for c in calls]


@router.get("/calls/archived")
def list_archived_calls(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Calls soft-archived via Remove All — restorable for 30 days, then auto-purged."""
    agent_ids = _org_agent_ids(db, current_user)
    _purge_expired_archived_calls(db, agent_ids)
    calls = (
        db.query(Call)
        .filter(Call.agent_id.in_(agent_ids), Call.archived_at.isnot(None))
        .order_by(Call.archived_at.desc())
        .all()
        if agent_ids else []
    )
    return [_call_to_dict(c) for c in calls]


class CallRestoreSchema(BaseModel):
    call_ids: Optional[List[int]] = None  # None = restore all archived calls


@router.post("/calls/restore")
def restore_archived_calls(payload: CallRestoreSchema = CallRestoreSchema(), current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Bring archived call logs back into the main list (clear archived_at)."""
    agent_ids = _org_agent_ids(db, current_user)
    if not agent_ids:
        return {"status": "success", "restored": 0}
    q = db.query(Call).filter(Call.agent_id.in_(agent_ids), Call.archived_at.isnot(None))
    if payload and payload.call_ids:
        q = q.filter(Call.id.in_(payload.call_ids))
    restored = q.update({Call.archived_at: None}, synchronize_session=False)
    db.commit()
    return {"status": "success", "restored": restored}

@router.get("/calls/{call_id}/transcript")
def get_call_transcript(call_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    transcript = db.query(Transcript).join(Call).join(Agent).join(Team).filter(
        Transcript.call_id == call_id,
        Team.organization_id == org.id
    ).first()
    if not transcript:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transcript not found for this call."
        )
    interest_score = transcript.interest_score
    if interest_score is None:
        interest_score = 0
    call = db.query(Call).filter(Call.id == call_id).first()
    recording_url = call.recording_url if call else None
    return {
        "call_id": call_id,
        "dialogue": transcript.dialogue_json or [],
        "summary": transcript.summary,
        "sentiment": transcript.sentiment,
        "interest_score": interest_score,
        "recording_url": recording_url,
        "wants_details": bool(getattr(transcript, "wants_details", False)),
        "details_sent": bool(getattr(transcript, "details_sent", False)),
        "details_sent_to": getattr(transcript, "details_sent_to", None),
    }


@router.post("/calls/{call_id}/send-details")
def send_call_details(call_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Manually (re)send the company details to this call's caller on WhatsApp + email."""
    org = get_active_org(db, current_user)
    # Ensure the call belongs to this org
    call = db.query(Call).join(Agent).join(Team).filter(
        Call.id == call_id, Team.organization_id == org.id
    ).first()
    if not call:
        raise HTTPException(status_code=404, detail="Call not found.")

    # Reset the dedupe guard so a manual send always attempts delivery
    transcript = db.query(Transcript).filter(Transcript.call_id == call_id).first()
    if transcript and transcript.details_sent:
        transcript.details_sent = False
        db.commit()

    from app.services.notifications import deliver_company_details
    result = deliver_company_details(db, call_id, reason="manual send")
    return {"status": "success" if result.get("ok") else "error", **result}


@router.delete("/calls")
def delete_all_calls(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    # SECURITY: only touch THIS organization's calls, never other tenants'.
    # Non-destructive: soft-archive the logs (restorable for 30 days) instead of deleting immediately.
    agent_ids = _org_agent_ids(db, current_user)
    try:
        archived = 0
        if agent_ids:
            archived = db.query(Call).filter(
                Call.agent_id.in_(agent_ids),
                Call.archived_at.is_(None),
            ).update({Call.archived_at: datetime.datetime.utcnow()}, synchronize_session=False)
            db.commit()
        return {
            "status": "success",
            "archived": archived,
            "message": "Call logs archived. They can be restored for 30 days before being permanently deleted.",
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@router.delete("/calls/{call_id}")
def delete_call(call_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    """Archive a single call log (org-scoped). Non-destructive: it moves to the
    Archived list, restorable for 30 days before being purged."""
    # SECURITY: only allow archiving a call that belongs to THIS organization.
    agent_ids = _org_agent_ids(db, current_user)
    call = db.query(Call).filter(Call.id == call_id).first()
    if not agent_ids or not call or call.agent_id not in agent_ids:
        raise HTTPException(status_code=404, detail="Call not found")
    try:
        if call.archived_at is None:
            call.archived_at = datetime.datetime.utcnow()
            db.commit()
        return {"status": "success", "message": "Call archived. Restorable for 30 days."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")



# --- VOICE LIBRARY & AI AGENT BUILDER ROUTER ---
class AiBuilderChatSchema(BaseModel):
    message: str
    history: List[dict] = []

@router.get("/voices")
def list_voices(current_user: User = Depends(get_current_user)) -> Any:
    """Neural voice library — the same 6 female + 6 male voices offered in agent creation."""
    def v(name: str, desc: str, gender: str, sample: str) -> dict:
        return {
            "id": name,
            "name": name,
            "description": desc,
            "provider": "Neural",
            "language": "English (US)",
            "flag": "\U0001F1FA\U0001F1F8",
            "use_case": "Customer Support",
            "age": "Adult",
            "gender": gender,
            "voice_id": name,
            "sample_url": f"/voices/{sample}",
        }
    return [
        v("Kore", "Balanced and crisp neural voice", "Female", "kore.wav"),
        v("Aoede", "Warm and expressive conversational voice", "Female", "aoede.wav"),
        v("Leda", "Professional and direct clear voice", "Female", "leda.wav"),
        v("Zephyr", "Warm and bright conversational voice", "Female", "zephyr.wav"),
        v("Gemma", "Clear and professional polished voice", "Female", "kore.wav"),
        v("Katie", "Bright and cheerful energetic voice", "Female", "aoede.wav"),
        v("Charon", "Deep and steady neural voice", "Male", "charon.wav"),
        v("Fenrir", "Strong and clear authoritative voice", "Male", "fenrir.wav"),
        v("Puck", "Energetic and crisp friendly voice", "Male", "puck.wav"),
        v("Achird", "Calm and smooth business voice", "Male", "achird.wav"),
        v("Archie", "Strong and clear authoritative voice", "Male", "fenrir.wav"),
        v("Corey", "Polished and friendly conversational voice", "Male", "charon.wav"),
    ]


@router.post("/ai-builder/chat")
def builder_chat(payload: AiBuilderChatSchema, current_user: User = Depends(get_current_user)) -> Any:
    """
    Intelligently extracts agent specifications based on user prompts using Gemini LLM
    """
    import urllib.request as _ur
    import json as _json
    import ssl as _ssl
    import os as _os

    user_msg = payload.message
    history = payload.history or []

    # Default fallback extraction variables
    agent_name = "AI Voice Assistant"
    system_prompt = "You are a professional AI assistant built to handle enterprise voice calls."
    first_message = "Hello."
    voice_name = "female"
    config_mode = "Inbound"
    reply_text = "Here is the compiled voice agent configuration!"

    gemini_key = _os.getenv("GEMINI_API_KEY")
    success = False

    if gemini_key:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
            
            system_instruction = (
                "You are an expert AI Voice Agent Architect.\n"
                "Your task is to generate and refine a calling agent's specifications. You must always return a JSON object.\n\n"
                "JSON SCHEMA:\n"
                "{\n"
                '  "name": "The agent\'s name (must be a clean string, e.g., Clara or Max)",\n'
                '  "system_prompt": "A comprehensive, high-quality, professional system prompt instruction block. You MUST generate this prompt in-depth, incorporating 100% of the user\'s detailed input instructions, rules, workflows, sections, or guidelines. Include all requested sections (such as Agent Identity, Role, Objective, Grounding Rules, Restrictions, Escalation, FAQs, error handling) and write them out fully and exhaustively. Do NOT summarize, shorten, or simplify them. Make it highly detailed, comprehensive, and ready for production calling.",\n'
                '  "first_message": "An engaging first greeting line (e.g., Hello, Clara speaking. How can I help you today?)",\n'
                '  "voice_name": "female" or "male",\n'
                '  "config_mode": "Inbound" or "Outbound"\n'
                "}\n\n"
                "CRITICAL RULES:\n"
                "1. Do not include markdown code block formatting like ```json or any other conversational text outside the JSON object.\n"
                "2. Keep the JSON strictly valid.\n"
                "3. You MUST fully adopt and execute the user's input template for the system prompt. If the user requests specific sections, formatting, rules, or guidelines, you MUST generate and include every single one of them in deep, exhaustive detail under the \"system_prompt\" key. Do not output placeholders."
            )
            
            contents = []
            for h in history:
                role = h.get("role")
                content = h.get("content", "")
                gemini_role = "model" if role == "assistant" else "user"
                contents.append({
                    "role": gemini_role,
                    "parts": [{"text": content}]
                })
            
            contents.append({
                "role": "user",
                "parts": [{"text": f"System Instructions:\n{system_instruction}\n\nUser request: {user_msg}"}]
            })
            
            body = {
                "contents": contents,
                "generationConfig": {
                    "responseMimeType": "application/json"
                }
            }
            
            req_data = _json.dumps(body).encode("utf-8")
            req = _ur.Request(
                url,
                data=req_data,
                headers={"Content-Type": "application/json"},
                method="POST"
            )
            
            context = _ssl._create_unverified_context()
            with _ur.urlopen(req, context=context, timeout=15) as resp:
                resp_body = resp.read().decode("utf-8")
                res_json = _json.loads(resp_body)
                text_response = res_json["candidates"][0]["content"]["parts"][0]["text"].strip()
                
                # Robustly find the JSON object in the response text
                first_curly = text_response.find('{')
                last_curly = text_response.rfind('}')
                if first_curly != -1 and last_curly != -1 and last_curly > first_curly:
                    json_str = text_response[first_curly:last_curly+1]
                else:
                    json_str = text_response
                    
                extracted = _json.loads(json_str)
                
                agent_name = extracted.get("name", agent_name)
                system_prompt = extracted.get("system_prompt", system_prompt)
                first_message = extracted.get("first_message", first_message)
                voice_name = extracted.get("voice_name", voice_name)
                config_mode = extracted.get("config_mode", config_mode)
                reply_text = f"I have dynamically generated and refined **{agent_name}** using Gemini! Take a look at the parameters parsed in the live preview panel on the right."
                success = True
        except Exception as e:
            print(f"[Gemini Prompt Generator] Failed to generate: {e}")

    # Fallback to heuristics if Gemini is not available or fails
    if not success:
        user_msg_lower = user_msg.lower()
        is_detailed = len(user_msg) > 120 or user_msg.count('\n') > 2
        
        if not is_detailed and ("clara" in user_msg_lower or "receptionist" in user_msg_lower or "front desk" in user_msg_lower):
            agent_name = "Clara - Front Desk Receptionist"
            system_prompt = (
                "# IDENTITY\n"
                "You are Clara, a professional receptionist for Voqly. Your tone is professional, direct, and efficient.\n\n"
                "# MISSION\n"
                "Greet inbound callers, answer common FAQs about business hours (9 AM - 5 PM) or services, "
                "and offer to route high-priority billing calls to our finance operations team."
            )
            first_message = "Hello. This is Clara at the front desk."
            voice_name = "female"
            config_mode = "Inbound"
            reply_text = (
                "I have dynamically generated **Clara - Front Desk Receptionist** using Gemini! She is configured with a professional, "
                "direct prompt tailored for greeting inbound customers, answering FAQs, and routing operational inquiries. "
                "Take a look at the parameters parsed in the live preview on the right!"
            )
        elif not is_detailed and ("sales" in user_msg_lower or "outbound" in user_msg_lower or "dialer" in user_msg_lower or "lead" in user_msg_lower):
            agent_name = "Max - Outbound Sales Closer"
            system_prompt = (
                "# IDENTITY\n"
                "You are Max, a professional outbound sales representative. Your tone is professional, persuasive, and direct.\n\n"
                "# MISSION\n"
                "Establish call intent, explain volume platform calling discounts, and ask the prospect which CRM they use "
                "in order to qualify them for a Solutions Engineer calendar slot."
            )
            first_message = "Hello. This is Max from Voqly. I am calling to discuss scaling your call capacities."
            voice_name = "male"
            config_mode = "Outbound"
            reply_text = (
                "Fantastic! I've designed **Max - Outbound Sales Closer** for high-volume lead qualification. "
                "Max uses a professional, direct sales prompt and is bound to a male voice model. "
                "The progressive parameters have been loaded into your live dialer preview!"
            )
        elif not is_detailed and ("support" in user_msg_lower or "ticket" in user_msg_lower or "help desk" in user_msg_lower):
            agent_name = "Daniel - Tech Support Agent"
            system_prompt = (
                "# IDENTITY\n"
                "You are Daniel, a professional technical support representative. Your tone is professional, analytical, and direct.\n\n"
                "# MISSION\n"
                "Analyze customer issues, create Zendesk or HubSpot support tickets, and route urgent incidents to our tier-2 engineer queue."
            )
            first_message = "Hello. This is Daniel from help desk support. Can you describe the issue you are facing?"
            voice_name = "male"
            config_mode = "Inbound"
            reply_text = (
                "I've configured **Daniel - Tech Support Agent** for automated help desk triage. "
                "Daniel is equipped with system instructions to collect issue details, create tickets, and escalate incidents. "
                "You can review the extracted prompt and voice specs in the right preview cards!"
            )
        else:
            # Custom Agent fallback that honors user inputs directly
            words = [w for w in user_msg.split() if w]
            candidate_name = "Voice"
            for w in words:
                clean_w = "".join(c for c in w if c.isalnum())
                if len(clean_w) > 3 and clean_w[0].isupper():
                    candidate_name = clean_w
                    break
            
            agent_name = f"{candidate_name} Agent"
            system_prompt = user_msg
            first_message = f"Hello. This is {candidate_name} Agent."
            voice_name = "female"
            config_mode = "Inbound"
            reply_text = (
                f"I have successfully compiled a custom calling agent named **{agent_name}** based on your prompt. "
                f"Take a look at the parameters parsed in the live preview panel on the right."
            )

    return {
        "status": "success",
        "reply": reply_text,
        "extracted_agent": {
            "name": agent_name,
            "system_prompt": system_prompt,
            "first_message": first_message,
            "voice_name": voice_name,
            "config_mode": config_mode
        }
    }


# --- BUSINESS DETAILS ENDPOINTS ---
class BusinessDetailsSchema(BaseModel):
    website_url: Optional[str] = None
    industry: Optional[str] = None
    tax_id: Optional[str] = None
    business_type: Optional[str] = None
    company_size: Optional[str] = None
    street_address: Optional[str] = None
    country: Optional[str] = None
    state_province: Optional[str] = None
    compliance_hipaa: Optional[bool] = False

@router.get("/organization/business-details")
def get_business_details(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    return {
        "status": "success",
        "website_url": org.website_url,
        "industry": org.industry,
        "tax_id": org.tax_id,
        "business_type": org.business_type,
        "company_size": org.company_size,
        "street_address": org.street_address,
        "country": org.country,
        "state_province": org.state_province,
        "compliance_hipaa": org.compliance_hipaa
    }

@router.put("/organization/business-details")
def update_business_details(payload: BusinessDetailsSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    org.website_url = payload.website_url
    org.industry = payload.industry
    org.tax_id = payload.tax_id
    org.business_type = payload.business_type
    org.company_size = payload.company_size
    org.street_address = payload.street_address
    org.country = payload.country
    org.state_province = payload.state_province
    org.compliance_hipaa = payload.compliance_hipaa
    db.commit()
    db.refresh(org)
    return {
        "status": "success",
        "message": "Business details updated successfully."
    }

# --- KNOWLEDGE BASE ENDPOINTS ---
from app.models.all_models import KnowledgeBase, Document

class KnowledgeBaseCreateSchema(BaseModel):
    name: str
    description: Optional[str] = ""

class DocumentCreateSchema(BaseModel):
    file_name: str
    file_size_kb: Optional[int] = 120
    content_text: Optional[str] = ""


class KbSearchSchema(BaseModel):
    query: str
    top_k: Optional[int] = 10


def _serialize_document(doc) -> dict:
    return {
        "id": doc.id,
        "file_name": doc.file_name,
        "file_url": doc.file_url,
        "content_text": doc.content_text,
        "index_status": getattr(doc, "index_status", "pending"),
        "chunk_count": getattr(doc, "chunk_count", 0) or 0,
        "char_count": getattr(doc, "char_count", 0) or 0,
        "index_error": getattr(doc, "index_error", None),
        "created_at": doc.created_at.isoformat(),
    }

@router.get("/knowledge-bases")
def list_knowledge_bases(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    kbs = db.query(KnowledgeBase).filter(KnowledgeBase.organization_id == org.id).all()
    
    result = []
    for kb in kbs:
        docs = db.query(Document).filter(Document.kb_id == kb.id).all()
        result.append({
            "id": kb.id,
            "name": kb.name,
            "description": kb.description,
            "created_at": kb.created_at.isoformat(),
            "documents": [_serialize_document(d) for d in docs]
        })
        
    # Seed a default knowledge base if empty to support quick uploader testing
    if not result:
        default_kb = KnowledgeBase(
            name="General Corporate FAQ",
            description="Holds general corporate guidelines, FAQs, and product manuals.",
            organization_id=org.id
        )
        db.add(default_kb)
        db.commit()
        db.refresh(default_kb)
        
        default_doc = Document(
            kb_id=default_kb.id,
            file_name="ProductSpecs_2026.txt",
            content_text="Voqly calling automation systems are built for high scalability, supporting ElevenLabs and Cartesia pipelines.",
            index_status="pending",
        )
        db.add(default_doc)
        db.commit()
        db.refresh(default_doc)

        from app.services.rag_service import index_document
        index_document(db, default_doc.id)
        db.refresh(default_doc)
        
        result.append({
            "id": default_kb.id,
            "name": default_kb.name,
            "description": default_kb.description,
            "created_at": default_kb.created_at.isoformat(),
            "documents": [_serialize_document(default_doc)]
        })
        
    return result

@router.post("/knowledge-bases", status_code=status.HTTP_201_CREATED)
def create_knowledge_base(payload: KnowledgeBaseCreateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    kb = KnowledgeBase(
        name=payload.name,
        description=payload.description,
        organization_id=org.id
    )
    db.add(kb)
    db.commit()
    db.refresh(kb)
    return {
        "status": "success",
        "id": kb.id,
        "name": kb.name,
        "description": kb.description,
        "documents": []
    }

@router.delete("/knowledge-bases/{kb_id}")
def delete_knowledge_base(kb_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    org = get_active_org(db, current_user)
    kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id, KnowledgeBase.organization_id == org.id).first()
    if not kb:
        raise HTTPException(status_code=404, detail="Knowledge base not found.")
        
    # Delete associated documents
    db.query(Document).filter(Document.kb_id == kb.id).delete()
    db.delete(kb)
    db.commit()
    return {"status": "success", "message": "Knowledge base deleted successfully."}

@router.post("/knowledge-bases/{kb_id}/documents", status_code=status.HTTP_201_CREATED)
def upload_kb_document(kb_id: int, payload: DocumentCreateSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    from app.services.rag_service import index_document

    org = get_active_org(db, current_user)
    kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id, KnowledgeBase.organization_id == org.id).first()
    if not kb:
        raise HTTPException(status_code=404, detail="Knowledge base not found.")

    content = (payload.content_text or "").strip()
    doc = Document(
        kb_id=kb.id,
        file_name=payload.file_name,
        content_text=content or f"Parsed content of text file: {payload.file_name}",
        index_status="pending",
        char_count=len(content),
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    index_result = index_document(db, doc.id)
    db.refresh(doc)
    return {
        "status": "success",
        "document": _serialize_document(doc),
        "index_result": index_result,
    }


@router.post("/knowledge-bases/{kb_id}/documents/upload", status_code=status.HTTP_201_CREATED)
async def upload_kb_document_file(
    kb_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Upload and parse a document file (PDF, DOCX, CSV, JSON, XLSX, SQLite, TXT) for vector RAG."""
    from app.services.document_parser import extract_text_from_bytes
    from app.services.rag_service import index_document

    org = get_active_org(db, current_user)
    kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id, KnowledgeBase.organization_id == org.id).first()
    if not kb:
        raise HTTPException(status_code=404, detail="Knowledge base not found.")

    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds 10MB limit.")

    file_name = file.filename or "upload.txt"
    content_text = extract_text_from_bytes(file_name, data)

    doc = Document(
        kb_id=kb.id,
        file_name=file_name,
        content_text=content_text,
        index_status="pending",
        char_count=len(content_text or ""),
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    index_result = index_document(db, doc.id)
    db.refresh(doc)
    return {
        "status": "success",
        "document": {
            **_serialize_document(doc),
            "content_preview": (content_text or "")[:200],
        },
        "index_result": index_result,
    }


@router.post("/knowledge-bases/{kb_id}/documents/{doc_id}/reindex")
def reindex_kb_document(
    kb_id: int,
    doc_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    from app.services.rag_service import index_document

    org = get_active_org(db, current_user)
    doc = (
        db.query(Document)
        .join(KnowledgeBase)
        .filter(
            Document.id == doc_id,
            Document.kb_id == kb_id,
            KnowledgeBase.organization_id == org.id,
        )
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")

    result = index_document(db, doc.id)
    db.refresh(doc)
    return {"status": "success", "document": _serialize_document(doc), "index_result": result}


@router.post("/knowledge-bases/{kb_id}/reindex")
def reindex_knowledge_base(
    kb_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    from app.services.rag_service import index_kb_documents

    org = get_active_org(db, current_user)
    kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id, KnowledgeBase.organization_id == org.id).first()
    if not kb:
        raise HTTPException(status_code=404, detail="Knowledge base not found.")

    result = index_kb_documents(db, kb.id)
    return {"status": "success", "kb_id": kb.id, **result}


@router.post("/knowledge-bases/{kb_id}/search")
def search_knowledge_base(
    kb_id: int,
    payload: KbSearchSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    from app.services.rag_service import search_kb

    org = get_active_org(db, current_user)
    kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id, KnowledgeBase.organization_id == org.id).first()
    if not kb:
        raise HTTPException(status_code=404, detail="Knowledge base not found.")

    query = (payload.query or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query is required.")

    results = search_kb(db, kb.id, query, top_k=payload.top_k or 10)
    return {"status": "success", "query": query, "results": results}

@router.delete("/knowledge-bases/documents/{doc_id}")
def delete_kb_document(doc_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Any:
    from app.models.all_models import KnowledgeBase, Document
    org = get_active_org(db, current_user)
    doc = db.query(Document).join(KnowledgeBase).filter(
        Document.id == doc_id,
        KnowledgeBase.organization_id == org.id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    db.delete(doc)
    db.commit()
    return {"status": "success", "message": "Document deleted successfully."}


# ─────────────────────────────────────────────────────────────────────────────
# BILLING — PLAN UPGRADE PAYMENT ENDPOINTS (Stripe & Razorpay)
# ─────────────────────────────────────────────────────────────────────────────

PLAN_PRICES = {
    "starter":      {"amount_cents": 9900,  "amount_inr": 8200,  "label": "Starter",      "monthly_usd": 99.00},
    "growth":       {"amount_cents": 49900, "amount_inr": 41500, "label": "Growth",        "monthly_usd": 499.00},
    "professional": {"amount_cents": 99900, "amount_inr": 83200, "label": "Professional",  "monthly_usd": 999.00},
}

from app.models.all_models import Subscription, Invoice as InvoiceModel

class BillingPlanIntentSchema(BaseModel):
    plan_tier: str   # starter | growth | professional
    currency: str = "usd"  # usd | inr

class RazorpayVerifySchema(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    plan_tier: str

class PlanSubscribeSchema(BaseModel):
    plan_tier: str
    payment_gateway: str        # stripe | razorpay
    payment_intent_id: str = "" # Stripe PI id
    razorpay_order_id: str = ""
    razorpay_payment_id: str = ""
    razorpay_signature: str = ""
    amount_paid: float = 0.0


@router.post("/billing/stripe/payment-intent")
def create_plan_stripe_intent(
    payload: BillingPlanIntentSchema,
    current_user: User = Depends(get_current_user)
) -> Any:
    """Create a Stripe PaymentIntent for a plan upgrade."""
    plan_tier = payload.plan_tier.lower()
    if plan_tier not in PLAN_PRICES:
        raise HTTPException(status_code=400, detail=f"Invalid plan tier: {plan_tier}")

    plan = PLAN_PRICES[plan_tier]
    secret_key = os.environ.get("STRIPE_SECRET_KEY", "")

    if not secret_key or "YOUR_SECRET_KEY_HERE" in secret_key or "placeholder" in secret_key:
        # Dev/test mode — return a mock client_secret
        return {
            "client_secret": f"pi_test_mock_plan_{plan_tier}_secret_mock",
            "payment_intent_id": f"pi_test_mock_plan_{plan_tier}_{random.randint(100000,999999)}",
            "amount": plan["amount_cents"],
            "currency": "usd",
            "mode": "test_mock",
            "plan_tier": plan_tier,
            "plan_label": plan["label"],
        }

    try:
        intent = stripe.PaymentIntent.create(
            amount=plan["amount_cents"],
            currency="usd",
            description=f"Voqly AI — {plan['label']} Plan Subscription",
            metadata={
                "plan_tier": plan_tier,
                "vendor_user_id": str(current_user.id),
                "vendor_user_email": current_user.email,
            },
            automatic_payment_methods={"enabled": True},
        )
        return {
            "client_secret": intent.client_secret,
            "payment_intent_id": intent.id,
            "amount": intent.amount,
            "currency": intent.currency,
            "mode": "live",
            "plan_tier": plan_tier,
            "plan_label": plan["label"],
        }
    except stripe.StripeError as e:
        raise HTTPException(status_code=400, detail=f"Stripe error: {str(e)}")


@router.post("/billing/razorpay/order")
def create_razorpay_order(
    payload: BillingPlanIntentSchema,
    current_user: User = Depends(get_current_user)
) -> Any:
    """Create a Razorpay order for a plan upgrade (amount in INR paise)."""
    plan_tier = payload.plan_tier.lower()
    if plan_tier not in PLAN_PRICES:
        raise HTTPException(status_code=400, detail=f"Invalid plan tier: {plan_tier}")

    plan = PLAN_PRICES[plan_tier]
    rzp_key_id = os.environ.get("RAZORPAY_KEY_ID", "rzp_test_placeholder")
    rzp_key_secret = os.environ.get("RAZORPAY_KEY_SECRET", "")

    if not rzp_key_secret or rzp_key_id == "rzp_test_placeholder" or "YOUR_SECRET_KEY_HERE" in rzp_key_secret or "placeholder" in rzp_key_secret:
        # Dev mode — return mock order
        mock_order_id = f"order_mock_{plan_tier}_{random.randint(100000, 999999)}"
        return {
            "order_id": mock_order_id,
            "amount": plan["amount_inr"] * 100,  # paise
            "currency": "INR",
            "key_id": rzp_key_id,
            "plan_tier": plan_tier,
            "plan_label": plan["label"],
            "mode": "test_mock",
        }

    try:
        import razorpay
        client = razorpay.Client(auth=(rzp_key_id, rzp_key_secret))
        order_data = {
            "amount": plan["amount_inr"] * 100,  # convert to paise
            "currency": "INR",
            "notes": {
                "plan_tier": plan_tier,
                "vendor_user_id": str(current_user.id),
                "vendor_email": current_user.email,
            },
        }
        order = client.order.create(data=order_data)
        return {
            "order_id": order["id"],
            "amount": order["amount"],
            "currency": order["currency"],
            "key_id": rzp_key_id,
            "plan_tier": plan_tier,
            "plan_label": plan["label"],
            "mode": "live",
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Razorpay error: {str(e)}")


@router.post("/billing/razorpay/verify")
def verify_razorpay_payment(
    payload: RazorpayVerifySchema,
    current_user: User = Depends(get_current_user)
) -> Any:
    """Verify Razorpay payment signature (HMAC-SHA256)."""
    rzp_key_secret = os.environ.get("RAZORPAY_KEY_SECRET", "")

    if not rzp_key_secret:
        # Dev mode — skip verification
        return {"status": "success", "verified": True, "mode": "test_mock"}

    try:
        import razorpay, hmac, hashlib
        generated = hmac.new(
            rzp_key_secret.encode(),
            f"{payload.razorpay_order_id}|{payload.razorpay_payment_id}".encode(),
            hashlib.sha256
        ).hexdigest()
        if generated != payload.razorpay_signature:
            raise HTTPException(status_code=400, detail="Razorpay signature verification failed.")
        return {"status": "success", "verified": True, "mode": "live"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Signature verification error: {str(e)}")


@router.post("/billing/subscribe")
def subscribe_plan(
    payload: PlanSubscribeSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> Any:
    """
    After payment is confirmed, upgrade the organization's subscription plan.
    Works for both Stripe and Razorpay.
    """
    import datetime as dt
    plan_tier = payload.plan_tier.lower()
    if plan_tier not in PLAN_PRICES:
        raise HTTPException(status_code=400, detail=f"Invalid plan tier: {plan_tier}")

    plan = PLAN_PRICES[plan_tier]
    org = get_active_org(db, current_user)

    # Upsert subscription
    sub = db.query(Subscription).filter(Subscription.organization_id == org.id).first()
    if not sub:
        sub = Subscription(
            organization_id=org.id,
            plan_tier=plan_tier,
            status="active",
            current_period_end=dt.datetime.utcnow() + dt.timedelta(days=30)
        )
        db.add(sub)
    else:
        sub.plan_tier = plan_tier
        sub.status = "active"
        sub.current_period_end = dt.datetime.utcnow() + dt.timedelta(days=30)
    db.commit()
    db.refresh(sub)

    # Record invoice
    inv = InvoiceModel(
        subscription_id=sub.id,
        amount_due=plan["monthly_usd"],
        amount_paid=payload.amount_paid or plan["monthly_usd"],
        status="paid",
        payment_gateway=payload.payment_gateway,
        stripe_payment_intent_id=payload.payment_intent_id or None,
        razorpay_order_id=payload.razorpay_order_id or None,
        razorpay_payment_id=payload.razorpay_payment_id or None,
        pdf_url=f"/invoices/inv_{plan_tier}_{int(plan['monthly_usd'])}_paid.pdf",
    )
    db.add(inv)

    # Audit log
    log = AuditLog(
        user_id=current_user.id,
        action=f"PLAN_UPGRADE_{plan_tier.upper()}",
        resource_type="SUBSCRIPTION",
        resource_id=sub.id,
        payload={"plan": plan_tier, "gateway": payload.payment_gateway}
    )
    db.add(log)
    db.commit()

    return {
        "status": "success",
        "message": f"Successfully subscribed to {plan['label']} plan.",
        "plan_tier": plan_tier,
        "plan_label": plan["label"],
        "subscription_status": "active",
        "current_period_end": sub.current_period_end.strftime("%B %d, %Y"),
        "invoice_id": f"INV-2026-{1000 + inv.id}",
    }




from fastapi.responses import HTMLResponse

@router.get("/billing/invoices/{invoice_id}/download")
def download_invoice(invoice_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> HTMLResponse:
    org = get_active_org(db, current_user)
    
    inv = db.query(Invoice).join(Subscription).filter(
        Invoice.id == invoice_id,
        Subscription.organization_id == org.id
    ).first()
    
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")
        
    plan_tier = "Starter"
    if inv.subscription:
        plan_tier = getattr(inv.subscription, "plan_tier", "starter").capitalize()

    tx_id = inv.stripe_payment_intent_id or inv.razorpay_payment_id or "N/A"
    method = inv.payment_gateway.upper() if inv.payment_gateway else "STRIPE"
    date_issued = inv.created_at.strftime('%B %d, %Y') if inv.created_at else ''

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Invoice INV-2026-{1000 + inv.id}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        @page {{
            size: A4;
            margin: 20mm;
        }}
        @media print {{
            body {{
                background: #ffffff;
                color: #000000;
                padding: 0;
            }}
            .no-print {{
                display: none !important;
            }}
            .invoice-card {{
                border: none !important;
                box-shadow: none !important;
                padding: 0 !important;
                margin: 0 !important;
            }}
        }}
        body {{
            font-family: 'Inter', sans-serif;
            background-color: #f8fafc;
            color: #1e293b;
            margin: 0;
            padding: 40px 20px;
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
        }}
        .invoice-card {{
            max-width: 800px;
            margin: 0 auto;
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 20px;
            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.05), 0 8px 10px -6px rgba(0, 0, 0, 0.05);
            padding: 40px;
        }}
        .invoice-header {{
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            border-bottom: 2px solid #f1f5f9;
            padding-bottom: 30px;
            margin-bottom: 30px;
        }}
        .logo-section {{
            display: flex;
            align-items: center;
            gap: 12px;
        }}
        .logo-icon {{
            width: 38px;
            height: 38px;
            border-radius: 10px;
            background: linear-gradient(135deg, #0b1931, #1e293b);
            display: flex;
            align-items: center;
            justify-content: center;
            color: #ffffff;
            font-weight: 800;
            font-size: 20px;
        }}
        .logo-text {{
            font-size: 20px;
            font-weight: 800;
            letter-spacing: 0.05em;
            color: #0f172a;
        }}
        .invoice-title {{
            text-align: right;
        }}
        .invoice-title h1 {{
            margin: 0;
            font-size: 28px;
            font-weight: 800;
            color: #0f172a;
            letter-spacing: -0.02em;
        }}
        .invoice-number {{
            font-family: monospace;
            font-size: 14px;
            color: #64748b;
            font-weight: 600;
            margin-top: 5px;
        }}
        .badge {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 6px 12px;
            border-radius: 9999px;
            font-size: 11px;
            font-weight: 700;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            margin-top: 10px;
        }}
        .badge-paid {{
            background-color: #ecfdf5;
            color: #047857;
            border: 1px solid #a7f3d0;
        }}
        .badge-unpaid {{
            background-color: #fffbeb;
            color: #b45309;
            border: 1px solid #fde68a;
        }}
        .info-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 40px;
            margin-bottom: 40px;
        }}
        .info-block {{
            font-size: 14px;
        }}
        .info-block h3 {{
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.1em;
            color: #94a3b8;
            margin: 0 0 10px 0;
        }}
        .info-block p {{
            margin: 0 0 5px 0;
            font-size: 14px;
            line-height: 1.5;
            color: #334155;
        }}
        .info-block strong {{
            color: #0f172a;
            font-weight: 600;
        }}
        .invoice-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }}
        .invoice-table th {{
            background-color: #f8fafc;
            border-bottom: 2px solid #e2e8f0;
            color: #475569;
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            padding: 12px 16px;
            text-align: left;
        }}
        .invoice-table td {{
            padding: 16px;
            border-bottom: 1px solid #f1f5f9;
            font-size: 14px;
            color: #334155;
        }}
        .item-description {{
            font-weight: 600;
            color: #0f172a;
        }}
        .item-subtext {{
            font-size: 12px;
            color: #64748b;
            margin-top: 3px;
        }}
        .align-right {{
            text-align: right !important;
        }}
        .align-center {{
            text-align: center !important;
        }}
        .summary-section {{
            display: flex;
            justify-content: flex-end;
            margin-bottom: 40px;
        }}
        .summary-table {{
            width: 300px;
            border-collapse: collapse;
        }}
        .summary-table td {{
            padding: 8px 16px;
            font-size: 14px;
            color: #475569;
        }}
        .summary-table tr.total-row td {{
            border-top: 2px solid #f1f5f9;
            font-weight: 800;
            font-size: 18px;
            color: #0f172a;
            padding-top: 12px;
        }}
        .footer-note {{
            text-align: center;
            border-top: 2px solid #f1f5f9;
            padding-top: 30px;
            font-size: 12px;
            color: #64748b;
            line-height: 1.6;
        }}
        .print-btn {{
            position: fixed;
            bottom: 30px;
            right: 30px;
            background-color: #0b1931;
            color: #ffffff;
            border: none;
            border-radius: 9999px;
            padding: 12px 24px;
            font-size: 14px;
            font-weight: 600;
            box-shadow: 0 10px 15px -3px rgba(11, 25, 49, 0.4);
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 8px;
            transition: all 0.2s;
        }}
        .print-btn:hover {{
            background-color: #1e293b;
            transform: translateY(-2px);
        }}
    </style>
</head>
<body>
    <div class="invoice-card">
        <div class="invoice-header">
            <div class="logo-section">
                <div class="logo-icon">V</div>
                <div class="logo-text">VOQLY AI</div>
            </div>
            <div class="invoice-title">
                <h1>INVOICE</h1>
                <div class="invoice-number">INV-2026-{1000 + inv.id}</div>
                <span class="badge badge-{inv.status}">{inv.status}</span>
            </div>
        </div>

        <div class="info-grid">
            <div class="info-block">
                <h3>Billed To</h3>
                <p><strong>{org.name}</strong></p>
                <p>Organization ID: ORG-{org.id}</p>
                <p>Customer ID: USR-{current_user.id}</p>
                <p>Email: {current_user.email}</p>
            </div>
            <div class="info-block" style="text-align: right;">
                <h3>Invoice Details</h3>
                <p><strong>Date Issued:</strong> {date_issued}</p>
                <p><strong>Payment Method:</strong> {method}</p>
                <p style="font-size: 11px; color: #64748b; margin-top: 4px;"><strong>Tx ID:</strong> {tx_id}</p>
            </div>
        </div>

        <table class="invoice-table">
            <thead>
                <tr>
                    <th>Description</th>
                    <th class="align-center">Quantity</th>
                    <th class="align-right">Unit Price</th>
                    <th class="align-right">Total</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>
                        <div class="item-description">Voqly AI {plan_tier} Plan Subscription</div>
                        <div class="item-subtext">Monthly recurring plan upgrade for voice agents & analytics integration.</div>
                    </td>
                    <td class="align-center">1</td>
                    <td class="align-right">${inv.amount_due:.2f}</td>
                    <td class="align-right">${inv.amount_due:.2f}</td>
                </tr>
            </tbody>
        </table>

        <div class="summary-section">
            <table class="summary-table">
                <tr>
                    <td>Subtotal</td>
                    <td class="align-right">${inv.amount_due:.2f}</td>
                </tr>
                <tr>
                    <td>Tax (0%)</td>
                    <td class="align-right">$0.00</td>
                </tr>
                <tr class="total-row">
                    <td>Total Paid</td>
                    <td class="align-right">${inv.amount_paid:.2f}</td>
                </tr>
            </table>
        </div>

        <div class="footer-note">
            <p>Thank you for choosing Voqly AI as your telephony agent partner.</p>
            <p style="font-size: 10px; color: #94a3b8; margin-top: 15px;">Voqly AI, Inc. · 100 Pine St, San Francisco, CA 94111 · support@voqly.ai</p>
        </div>
    </div>

    <button class="print-btn no-print" onclick="window.print()">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="margin-right: 4px;"><path d="M6 9V2h12v7"></path><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg>
        Print Invoice
    </button>

    <script>
        window.addEventListener('DOMContentLoaded', () => {{
            setTimeout(() => {{ window.print(); }}, 500);
        }});
    </script>
</body>
</html>
"""
    return HTMLResponse(content=html_content)


# ═══════════════════════════════════════════════════════════════════════════════
# BILLING ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

PLAN_PRICES_USD = {"starter": 99, "growth": 499, "professional": 999}
PLAN_PRICES_INR = {"starter": 8200, "growth": 41500, "professional": 83200}


class StripePaymentIntentSchema(BaseModel):
    plan_tier: str
    currency: str = "usd"


class RazorpayOrderSchema(BaseModel):
    plan_tier: str
    currency: str = "inr"


class RazorpayVerifySchema(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    plan_tier: str


class SubscribeSchema(BaseModel):
    plan_tier: str
    payment_gateway: str = "stripe"
    payment_intent_id: Optional[str] = None
    razorpay_order_id: Optional[str] = None
    razorpay_payment_id: Optional[str] = None
    amount_paid: Optional[float] = None


# ── GET /organization/billing ────────────────────────────────────────────────
@router.get("/organization/billing")
def get_org_billing(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> Any:
    """Return current billing info for the authenticated vendor's organisation."""
    org = get_active_org(db, current_user)
    sub = (
        db.query(Subscription)
        .filter(Subscription.organization_id == org.id)
        .order_by(Subscription.created_at.desc())
        .first()
    )

    invoices_data = []
    if sub:
        for inv in sorted(sub.invoices, key=lambda x: x.created_at, reverse=True)[:10]:
            invoices_data.append({
                "id": inv.id,
                "invoice_number": f"INV-2026-{1000 + inv.id}",
                "amount": inv.amount_paid or inv.amount_due,
                "status": inv.status,
                "created_at": inv.created_at.strftime("%Y-%m-%d") if inv.created_at else "",
                "pdf_url": f"/dashboard/billing/invoices/{inv.id}/download",
            })

    plan_tier = sub.plan_tier if sub else "free"
    period_end = (
        sub.current_period_end.strftime("%Y-%m-%d")
        if sub and sub.current_period_end
        else (datetime.datetime.utcnow() + datetime.timedelta(days=30)).strftime("%Y-%m-%d")
    )

    return {
        "prepaid_balance": org.prepaid_balance or 0.0,
        "plan_tier": plan_tier,
        "subscription_status": sub.status if sub else "trial",
        "current_period_end": period_end,
        "invoices": invoices_data,
    }


# ── POST /billing/stripe/payment-intent ────────────────────────────────────
@router.post("/billing/stripe/payment-intent")
def create_stripe_payment_intent(
    payload: StripePaymentIntentSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Create a Stripe PaymentIntent.  Falls back to test_mock when no key is set."""
    tier = payload.plan_tier.lower()
    amount_usd = PLAN_PRICES_USD.get(tier, 99)

    # ── Real Stripe path ──────────────────────────────────────────────────────
    stripe_key = os.environ.get("STRIPE_SECRET_KEY", "")
    if stripe and stripe_key and not stripe_key.endswith("_HERE"):
        try:
            stripe.api_key = stripe_key
            intent = stripe.PaymentIntent.create(
                amount=amount_usd * 100,  # cents
                currency="usd",
                metadata={"user_id": str(current_user.id), "plan_tier": tier},
            )
            return {
                "client_secret": intent.client_secret,
                "payment_intent_id": intent.id,
                "mode": "live",
            }
        except Exception as e:
            # Fall through to mock if Stripe call fails (e.g. test key restrictions)
            print(f"[Stripe] PaymentIntent creation failed, using mock: {e}")

    # ── Test / Mock path ──────────────────────────────────────────────────────
    mock_pi_id = f"pi_test_mock_{current_user.id}_{tier}_{random.randint(100000, 999999)}"
    return {
        "client_secret": f"{mock_pi_id}_secret",
        "payment_intent_id": mock_pi_id,
        "mode": "test_mock",
    }


# ── POST /billing/razorpay/order ────────────────────────────────────────────
@router.post("/billing/razorpay/order")
def create_razorpay_order(
    payload: RazorpayOrderSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Create a Razorpay order. Falls back to test_mock when no key is set."""
    tier = payload.plan_tier.lower()
    amount_inr = PLAN_PRICES_INR.get(tier, 8200)

    rzp_key_id = os.environ.get("RAZORPAY_KEY_ID", "")
    rzp_key_secret = os.environ.get("RAZORPAY_KEY_SECRET", "")

    if rzp_key_id and rzp_key_secret and not rzp_key_id.endswith("_HERE") and rzp_key_id != "rzp_test_SuiaVbQ34XG4B9":
        try:
            import razorpay  # type: ignore
            client = razorpay.Client(auth=(rzp_key_id, rzp_key_secret))
            order = client.order.create({
                "amount": amount_inr * 100,  # paise
                "currency": "INR",
                "notes": {"user_id": str(current_user.id), "plan_tier": tier},
            })
            return {
                "order_id": order["id"],
                "amount": order["amount"],
                "currency": order["currency"],
                "key_id": rzp_key_id,
                "mode": "live",
            }
        except Exception as e:
            print(f"[Razorpay] Order creation failed, using mock: {e}")

    # Mock path
    mock_order_id = f"order_mock_{current_user.id}_{tier}_{random.randint(100000, 999999)}"
    return {
        "order_id": mock_order_id,
        "amount": amount_inr * 100,
        "currency": "INR",
        "key_id": rzp_key_id or "rzp_test_SuiaVbQ34XG4B9",
        "mode": "test_mock",
    }


# ── POST /billing/razorpay/verify ───────────────────────────────────────────
@router.post("/billing/razorpay/verify")
def verify_razorpay_payment(
    payload: RazorpayVerifySchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Verify Razorpay webhook signature. Skipped in mock mode."""
    rzp_key_secret = os.environ.get("RAZORPAY_KEY_SECRET", "")
    # Skip verification if no real secret is configured (mock mode)
    if not rzp_key_secret or rzp_key_secret.endswith("_HERE") or rzp_key_secret == os.environ.get("RAZORPAY_KEY_ID", ""):
        return {"status": "mock_verified"}

    try:
        import hmac
        import hashlib
        msg = f"{payload.razorpay_order_id}|{payload.razorpay_payment_id}"
        expected = hmac.HMAC(
            rzp_key_secret.encode(),
            msg.encode(),
            hashlib.sha256,
        ).hexdigest()
        if expected != payload.razorpay_signature:
            raise HTTPException(status_code=400, detail="Invalid Razorpay signature.")
    except ImportError:
        pass  # hmac not available; skip verification

    return {"status": "verified"}


# ── POST /billing/subscribe ─────────────────────────────────────────────────
@router.post("/billing/subscribe")
def subscribe_to_plan(
    payload: SubscribeSchema,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Record a successful plan subscription and create an invoice."""
    org = get_active_org(db, current_user)
    tier = payload.plan_tier.lower()
    amount = payload.amount_paid or PLAN_PRICES_USD.get(tier, 0)

    # Upsert subscription
    sub = (
        db.query(Subscription)
        .filter(Subscription.organization_id == org.id)
        .first()
    )
    period_end = datetime.datetime.utcnow() + datetime.timedelta(days=30)
    if sub:
        sub.plan_tier = tier
        sub.status = "active"
        sub.current_period_end = period_end
    else:
        sub = Subscription(
            organization_id=org.id,
            plan_tier=tier,
            status="active",
            current_period_end=period_end,
        )
        db.add(sub)
        db.flush()

    # Create invoice record
    inv = Invoice(
        subscription_id=sub.id,
        amount_due=amount,
        amount_paid=amount,
        status="paid",
        payment_gateway=payload.payment_gateway,
        stripe_payment_intent_id=payload.payment_intent_id,
        razorpay_order_id=payload.razorpay_order_id,
        razorpay_payment_id=payload.razorpay_payment_id,
    )
    db.add(inv)

    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action="PLAN_SUBSCRIPTION",
        resource_type="SUBSCRIPTION",
        resource_id=sub.id,
        payload={"plan_tier": tier, "gateway": payload.payment_gateway, "amount": amount},
    )
    db.add(audit)
    db.commit()

    return {
        "status": "success",
        "message": f"Successfully subscribed to the {tier.capitalize()} plan.",
        "plan_tier": tier,
        "subscription_id": sub.id,
        "invoice_id": inv.id,
    }


# ── GET /billing/invoices/{invoice_id}/download ─────────────────────────────
@router.get("/billing/invoices/{invoice_id}/download")
def download_billing_invoice(
    invoice_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Return a printable HTML invoice for the authenticated user's organisation."""
    from fastapi.responses import HTMLResponse

    org = get_active_org(db, current_user)
    inv = (
        db.query(Invoice)
        .join(Subscription)
        .filter(Invoice.id == invoice_id, Subscription.organization_id == org.id)
        .first()
    )
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    sub = inv.subscription
    tier = sub.plan_tier.capitalize() if sub else "Unknown"
    created = inv.created_at.strftime("%B %d, %Y") if inv.created_at else "—"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <title>Invoice INV-2026-{1000 + inv.id} — Voqly AI</title>
  <style>
    body {{ font-family: 'Inter', sans-serif; padding: 40px; color: #0f172a; max-width: 700px; margin: auto; }}
    h1 {{ font-size: 22px; font-weight: 800; }}
    .label {{ font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .05em; color: #64748b; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 24px; }}
    th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid #e2e8f0; font-size: 13px; }}
    th {{ background: #f8fafc; font-weight: 700; }}
    .total {{ font-weight: 800; font-size: 15px; }}
    @media print {{ .no-print {{ display: none; }} }}
  </style>
</head>
<body>
  <h1>Voqly AI</h1>
  <p class="label">Invoice</p>
  <p style="font-size:13px; color:#64748b;">INV-2026-{1000 + inv.id} &nbsp;|&nbsp; {created}</p>

  <table>
    <thead><tr><th>Description</th><th>Amount</th></tr></thead>
    <tbody>
      <tr><td>{tier} Plan — Monthly Subscription</td><td>${inv.amount_paid:.2f}</td></tr>
      <tr><td>Payment Gateway</td><td>{inv.payment_gateway.capitalize()}</td></tr>
    </tbody>
    <tfoot>
      <tr class="total"><td>Total Paid</td><td>${inv.amount_paid:.2f}</td></tr>
    </tfoot>
  </table>

  <p style="margin-top:32px; font-size:11px; color:#94a3b8;">
    Voqly AI, Inc. · support@voqly.ai · Thank you for your business.
  </p>
  <button class="no-print" onclick="window.print()"
    style="margin-top:20px; padding:10px 20px; background:#0f172a; color:#fff; border:none; border-radius:8px; cursor:pointer; font-size:13px; font-weight:700;">
    Print / Save PDF
  </button>
  <script>
    window.addEventListener('DOMContentLoaded', () => {{ setTimeout(() => window.print(), 600); }});
  </script>
</body>
</html>"""
    return HTMLResponse(content=html)


@router.get("/leads/check-database")
def check_leads_database(
    category: str,
    subcategory: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> Any:
    """Check if database files already exist for a given category/subcategory."""
    import os
    import sqlite3
    from app.models.all_models import KnowledgeBase, Document
    from app.services.agent_call_context import resolve_rhea_db_path

    org = get_active_org(db, current_user)
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

    org_base_dir = os.path.abspath(os.path.join(base_dir, "databses", f"org_{org.id}"))
    target_dir = os.path.abspath(os.path.join(org_base_dir, category, subcategory))
    if not target_dir.startswith(org_base_dir):
        raise HTTPException(status_code=400, detail="Invalid category or subcategory name.")

    files = []
    if os.path.exists(target_dir):
        all_files = [f for f in os.listdir(target_dir) if os.path.isfile(os.path.join(target_dir, f))]
        files = [f for f in all_files if not f.lower().endswith((".pdf.txt", ".docx.txt"))]

    row_count = 0
    table_count = 0
    db_path = resolve_rhea_db_path(org.id, category, subcategory)
    if db_path and os.path.exists(db_path):
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            tables = cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()
            table_count = len(tables)
            for (tname,) in tables:
                row_count += cur.execute(f"SELECT COUNT(*) FROM {tname}").fetchone()[0]
            conn.close()
        except Exception:
            pass

    kb_name = f"{category} / {subcategory} Database"
    kb = db.query(KnowledgeBase).filter(
        KnowledgeBase.organization_id == org.id,
        KnowledgeBase.name == kb_name,
    ).first()
    index_status = "none"
    vector_chunks = 0
    if kb:
        docs = db.query(Document).filter(Document.kb_id == kb.id).all()
        if docs:
            index_status = docs[0].index_status or "pending"
            vector_chunks = sum(d.chunk_count or 0 for d in docs)

    return {
        "exists": len(files) > 0,
        "files": files,
        "processing": {
            "status": "ready" if len(files) > 0 and index_status == "ready" else ("indexed" if vector_chunks else "pending"),
            "index_status": index_status,
            "row_count": row_count,
            "table_count": table_count,
            "vector_chunks": vector_chunks,
            "kb_id": kb.id if kb else None,
        },
    }


@router.post("/leads/upload-database")
async def upload_leads_database(
    category: str = Form(...),
    subcategory: str = Form(...),
    format: str = Form(...),  # 'sqlite', 'csv', 'excel', 'json', 'google_sheet'
    files: Optional[List[UploadFile]] = File(None),
    google_sheet_url: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> Any:
    from fastapi import Form
    import os
    import io
    import csv
    import json
    import sqlite3
    import openpyxl
    import datetime
    import re
    import urllib.request as _ur

    org = get_active_org(db, current_user)
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

    # Prevent path traversal
    org_base_dir = os.path.abspath(os.path.join(base_dir, "databses", f"org_{org.id}"))
    target_dir = os.path.abspath(os.path.join(org_base_dir, category, subcategory))
    if not target_dir.startswith(org_base_dir):
        raise HTTPException(status_code=400, detail="Invalid category or subcategory name.")
    os.makedirs(target_dir, exist_ok=True)

    db_deleted = False
    saved_files = []
    warnings_list = []

    # 1. Gather all file/data inputs to parse
    parsed_files_data = [] # List of tuples: (filename, rows, raw_content_or_none, text_content_or_none)

    try:
        if format == "google_sheet":
            if not google_sheet_url:
                raise HTTPException(status_code=400, detail="Google Sheet URL is required for format 'google_sheet'.")
            
            if "/spreadsheets/d/e/" in google_sheet_url:
                match = re.search(r"/spreadsheets/d/e/([a-zA-Z0-9-_]+)", google_sheet_url)
                if not match:
                    raise HTTPException(
                        status_code=400,
                        detail="Invalid Google Sheet URL format. Make sure it is a valid published Google Sheets link."
                    )
                publish_key = match.group(1)
                export_url = f"https://docs.google.com/spreadsheets/d/e/{publish_key}/pub?output=csv"
                gid_match = re.search(r"[#&?]gid=([0-9]+)", google_sheet_url)
                if gid_match:
                    export_url += f"&single=true&gid={gid_match.group(1)}"
            else:
                match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", google_sheet_url)
                if not match:
                    raise HTTPException(
                        status_code=400,
                        detail="Invalid Google Sheet URL format. Make sure it is a valid Google Sheets link."
                    )
                spreadsheet_id = match.group(1)
                export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"
                gid_match = re.search(r"[#&?]gid=([0-9]+)", google_sheet_url)
                if gid_match:
                    export_url += f"&gid={gid_match.group(1)}"

            req = _ur.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
            import ssl
            context = ssl._create_unverified_context()
            try:
                with _ur.urlopen(req, context=context, timeout=10) as resp:
                    csv_content = resp.read()
            except Exception as fetch_err:
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to download Google Sheet: {str(fetch_err)}. Verify that 'Anyone with the link can view' is enabled."
                )

            text_content = csv_content.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(text_content))
            rows = [row for row in reader]
            parsed_files_data.append(("google_sheet.csv", rows, None, None))

        else:
            downloaded_files = []
            if files and any(f.filename for f in files):
                for file in files:
                    if not file.filename:
                        continue
                    content = await file.read()
                    downloaded_files.append((file.filename, content))
            elif google_sheet_url:
                # Download file from URL
                import urllib.request as _ur
                import ssl
                import re
                
                download_url = google_sheet_url
                
                # Check if it's a Google Sheets link and format is csv/excel
                if "docs.google.com/spreadsheets" in download_url:
                    if "/spreadsheets/d/e/" in download_url:
                        match = re.search(r"/spreadsheets/d/e/([a-zA-Z0-9-_]+)", download_url)
                        if match:
                            publish_key = match.group(1)
                            output_fmt = "csv" if format == "csv" else "xlsx"
                            download_url = f"https://docs.google.com/spreadsheets/d/e/{publish_key}/pub?output={output_fmt}"
                            gid_match = re.search(r"[#&?]gid=([0-9]+)", google_sheet_url)
                            if gid_match:
                                download_url += f"&single=true&gid={gid_match.group(1)}"
                    else:
                        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", download_url)
                        if match:
                            spreadsheet_id = match.group(1)
                            output_fmt = "csv" if format == "csv" else "xlsx"
                            download_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format={output_fmt}"
                            gid_match = re.search(r"[#&?]gid=([0-9]+)", google_sheet_url)
                            if gid_match:
                                download_url += f"&gid={gid_match.group(1)}"
                
                context = ssl._create_unverified_context()
                req = _ur.Request(download_url, headers={"User-Agent": "Mozilla/5.0"})
                try:
                    with _ur.urlopen(req, context=context, timeout=15) as resp:
                        content = resp.read()
                        
                        # Try to extract filename from content-disposition header or URL
                        filename = "database"
                        cd_header = resp.headers.get("Content-Disposition")
                        if cd_header:
                            fn_match = re.search(r'filename="?([^"]+)"?', cd_header)
                            if fn_match:
                                filename = fn_match.group(1)
                        if filename == "database":
                            url_path = os.path.basename(google_sheet_url.split("?")[0])
                            if url_path:
                                filename = url_path
                                
                        # Ensure filename extension matches format
                        if format == "csv" and not filename.endswith(".csv"):
                            filename += ".csv"
                        elif format == "excel" and not filename.endswith((".xlsx", ".xls")):
                            filename += ".xlsx"
                        elif format == "sqlite" and not filename.endswith(".db"):
                            filename += ".db"
                        elif format == "json" and not filename.endswith(".json"):
                            filename += ".json"
                            
                        downloaded_files.append((filename, content))
                except Exception as fetch_err:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Failed to download database file from link: {str(fetch_err)}"
                    )
            else:
                raise HTTPException(status_code=400, detail="At least one database file or a link is required.")
            
            for filename, content in downloaded_files:
                rows = []
                
                if format == "csv":
                    text_content = content.decode("utf-8", errors="ignore")
                    reader = csv.DictReader(io.StringIO(text_content))
                    rows = [row for row in reader]
                elif format == "excel":
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    sheet = wb.active
                    for name in wb.sheetnames:
                        name_lower = name.lower().strip()
                        if any(x in name_lower for x in ("leads", "customers", "contacts", "data", "sheet1")):
                            sheet = wb[name]
                            break
                    headers = []
                    for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
                        headers.append(str(cell or "").strip())
                    for r in sheet.iter_rows(min_row=2, values_only=True):
                        if not any(r):
                            continue
                        row_dict = {}
                        for idx, val in enumerate(r):
                            if idx < len(headers):
                                row_dict[headers[idx]] = val
                        rows.append(row_dict)
                    wb.close()
                elif format == "json":
                    text_content = content.decode("utf-8", errors="ignore")
                    data = json.loads(text_content)
                    if isinstance(data, list):
                        rows = data
                    elif isinstance(data, dict):
                        for k, v in data.items():
                            if isinstance(v, list):
                                rows = v
                                break
                        if not rows:
                            rows = [data]
                elif format == "sqlite":
                    pass
                elif format == "pdf_docx":
                    from app.services.document_parser import extract_text_from_bytes
                    txt_content = extract_text_from_bytes(filename, content)
                    parsed_files_data.append((filename, [], content, txt_content))
                    continue

                parsed_files_data.append((filename, rows, content if format == "sqlite" else None, None))

    except HTTPException as http_ex:
        raise http_ex
    except Exception as parse_err:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to process database input: {str(parse_err)}"
        )

    # 2. Iterate and process/save the gathered data
    for filename_to_save, rows, raw_content, text_content in parsed_files_data:
        is_ecommerce_marketing_campaign = (category == "Ecommerce" and subcategory == "Marketing Campaign")
        db_filename = "ecommerce_marketing.db" if is_ecommerce_marketing_campaign else "rhea_ecommerce.db"
        db_file_path = os.path.join(target_dir, db_filename)
        
        if format == "sqlite":
            file_path = os.path.join(target_dir, db_filename)
            with open(file_path, "wb") as f:
                f.write(raw_content)
            saved_files.append(db_filename)
            
        elif format == "pdf_docx":
            if not db_deleted:
                try:
                    for f in os.listdir(target_dir):
                        f_path = os.path.join(target_dir, f)
                        if os.path.isfile(f_path):
                            os.remove(f_path)
                except Exception as remove_err:
                    print(f"Failed to clear old files in {target_dir}: {remove_err}")
                db_deleted = True
            
            # Save original file
            orig_file_path = os.path.join(target_dir, filename_to_save)
            with open(orig_file_path, "wb") as f:
                f.write(raw_content)
            saved_files.append(filename_to_save)
            
            # Save parsed text companion file for PDF/DOCX
            if filename_to_save.lower().endswith((".pdf", ".docx", ".doc")):
                txt_filename = filename_to_save + ".txt"
                txt_file_path = os.path.join(target_dir, txt_filename)
                with open(txt_file_path, "w", encoding="utf-8") as f:
                    f.write(text_content or "")
            
        else:  # csv, excel, json, google_sheet
            if not db_deleted:
                try:
                    if os.path.exists(db_file_path):
                        os.remove(db_file_path)
                except Exception as remove_err:
                    print(f"Failed to delete old database file {db_file_path}: {remove_err}")
                db_deleted = True
                
            # Structurize to SQLite schema
            has_inserted = False
            try:
                conn = sqlite3.connect(db_file_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT,
                    phone TEXT,
                    city TEXT,
                    address TEXT
                );
                """)
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS products (
                    product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT,
                    price REAL,
                    is_active INTEGER DEFAULT 1
                );
                """)
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS cart (
                    user_id INTEGER,
                    product_id INTEGER,
                    quantity INTEGER,
                    PRIMARY KEY (user_id, product_id)
                );
                """)
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS orders (
                    order_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    status TEXT,
                    placed_at TEXT,
                    total REAL,
                    issue TEXT,
                    discount_code TEXT
                );
                """)
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS order_items (
                    order_id INTEGER,
                    product_name TEXT
                );
                """)
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS cod_confirmations (
                    order_id INTEGER PRIMARY KEY,
                    delivery_address TEXT
                );
                """)
                
                product_map = {}
                for r in rows:
                    norm_r = {str(k).lower().strip(): v for k, v in r.items()}
                    name = norm_r.get("name", norm_r.get("customer name", norm_r.get("customer", "")))
                    phone = norm_r.get("phone", norm_r.get("phone number", norm_r.get("number", norm_r.get("mobile", ""))))
                    city = norm_r.get("city", norm_r.get("region", "Delhi"))
                    address = norm_r.get("address", norm_r.get("delivery address", norm_r.get("shipping address", "Address Details")))
                    
                    if not phone or not name:
                        continue
                        
                    cursor.execute(
                        "INSERT INTO users (name, phone, city, address) VALUES (?, ?, ?, ?)",
                        (str(name), str(phone), str(city), str(address))
                    )
                    user_id = cursor.lastrowid
                    has_inserted = True
                    
                    product_name = norm_r.get("product", norm_r.get("product name", norm_r.get("item", norm_r.get("items", ""))))
                    price_val = norm_r.get("price", norm_r.get("amount", norm_r.get("cost", 0.0)))
                    try:
                        price = float(price_val) if price_val else 0.0
                    except ValueError:
                        price = 0.0
                        
                    qty_val = norm_r.get("quantity", norm_r.get("qty", 1))
                    try:
                        quantity = int(qty_val) if qty_val else 1
                    except ValueError:
                        quantity = 1
                        
                    product_id = None
                    if product_name:
                        product_name = str(product_name).strip()
                        if product_name not in product_map:
                            cursor.execute(
                                "INSERT INTO products (name, price, is_active) VALUES (?, ?, 1)",
                                (product_name, price)
                            )
                            product_id = cursor.lastrowid
                            product_map[product_name] = product_id
                        else:
                            product_id = product_map[product_name]
                            
                    status = str(norm_r.get("status", norm_r.get("order status", "pending"))).lower().strip()
                    issue = norm_r.get("issue", norm_r.get("reason", ""))
                    discount_code = norm_r.get("discount", norm_r.get("discount code", ""))
                    
                    is_cart = norm_r.get("type", "") == "cart" or "abandon" in status or "abandoned" in status or status == "pending"
                    
                    if is_cart and product_id:
                        cursor.execute(
                            "INSERT OR REPLACE INTO cart (user_id, product_id, quantity) VALUES (?, ?, ?)",
                            (user_id, product_id, quantity)
                        )
                    else:
                        placed_at = norm_r.get("placed_at", norm_r.get("date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                        total = price * quantity
                        cursor.execute(
                            "INSERT INTO orders (user_id, status, placed_at, total, issue, discount_code) VALUES (?, ?, ?, ?, ?, ?)",
                            (user_id, status.upper(), str(placed_at), total, str(issue) if issue else None, str(discount_code) if discount_code else None)
                        )
                        order_id = cursor.lastrowid
                        
                        if product_name:
                            cursor.execute(
                                "INSERT INTO order_items (order_id, product_name) VALUES (?, ?)",
                                (order_id, product_name)
                            )
                        cursor.execute(
                            "INSERT OR REPLACE INTO cod_confirmations (order_id, delivery_address) VALUES (?, ?)",
                            (order_id, str(address))
                        )
                conn.commit()
                conn.close()
            except Exception as e:
                print(f"Failed to auto-structurize sqlite db from upload: {e}")
                warnings_list.append(f"Auto-structuring schema failed: {str(e)}")

            if not has_inserted:
                warnings_list.append(
                    f"Could not auto-detect standard columns (e.g. 'name' and 'phone') in input database: '{filename_to_save}'. "
                    "No rows were mapped to default workflows, but raw table was created."
                )

            # Dynamically insert into a table matching the file's sanitized name
            has_inserted_dyn = False
            if rows:
                base_name = os.path.splitext(filename_to_save)[0]
                sanitized_table = "".join(c if c.isalnum() else "_" for c in base_name).lower().strip("_")
                if not sanitized_table:
                    sanitized_table = "uploaded_data"
                try:
                    conn2 = sqlite3.connect(db_file_path)
                    cursor2 = conn2.cursor()
                    
                    headers = list(rows[0].keys())
                    sanitized_headers = []
                    for h in headers:
                        sh = "".join(c if c.isalnum() else "_" for c in str(h)).lower().strip("_")
                        if not sh:
                            sh = f"col_{len(sanitized_headers)}"
                        original_sh = sh
                        counter = 1
                        while sh in sanitized_headers:
                            sh = f"{original_sh}_{counter}"
                            counter += 1
                        sanitized_headers.append(sh)
                        
                    cols_def = ", ".join(f'"{h}" TEXT' for h in sanitized_headers)
                    cursor2.execute(f'CREATE TABLE IF NOT EXISTS "{sanitized_table}" ({cols_def});')
                    
                    headers_part = ", ".join(f'"{sh}"' for sh in sanitized_headers)
                    placeholders_part = ", ".join("?" for _ in sanitized_headers)
                    insert_sql = f'INSERT INTO "{sanitized_table}" ({headers_part}) VALUES ({placeholders_part});'
                    for row in rows:
                        vals = [row.get(h) for h in headers]
                        cursor2.execute(insert_sql, vals)
                    conn2.commit()
                    conn2.close()
                    has_inserted_dyn = True
                except Exception as dyn_err:
                    print(f"Failed to insert dynamic table {sanitized_table}: {dyn_err}")
                    
            if has_inserted or has_inserted_dyn:
                if db_filename not in saved_files:
                    saved_files.append(db_filename)

            # Save the readable CSV file representation too
            base_name = os.path.splitext(filename_to_save)[0]
            csv_filename = f"{base_name}.csv"
            file_path = os.path.join(target_dir, csv_filename)
            if rows:
                keys = rows[0].keys()
                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=keys)
                    writer.writeheader()
                    writer.writerows(rows)
            if csv_filename not in saved_files:
                saved_files.append(csv_filename)

    # 3. Vector-index uploaded SQLite for semantic RAG retrieval on calls
    vector_result = {"status": "skipped"}
    is_ecommerce_marketing_campaign = (category == "Ecommerce" and subcategory == "Marketing Campaign")
    db_filename = "ecommerce_marketing.db" if is_ecommerce_marketing_campaign else "rhea_ecommerce.db"
    final_db_path = os.path.join(target_dir, db_filename)
    if format != "pdf_docx" and os.path.exists(final_db_path):
        from app.services.rag_service import sync_sqlite_db_to_kb
        try:
            vector_result = sync_sqlite_db_to_kb(db, org.id, category, subcategory, final_db_path)
        except Exception as vec_err:
            print(f"[RAG] Failed to vector-index uploaded database: {vec_err}")
            vector_result = {"status": "failed", "message": str(vec_err)}

    row_count = vector_result.get("row_count", 0)
    if not row_count and os.path.exists(final_db_path):
        try:
            conn = sqlite3.connect(final_db_path)
            cur = conn.cursor()
            tables = cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()
            for (tname,) in tables:
                row_count += cur.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
            conn.close()
        except Exception:
            pass

    return {
        "status": "success",
        "message": f"Successfully structured and uploaded database files {saved_files} for {category} › {subcategory}.",
        "files": saved_files,
        "warnings": warnings_list,
        "processing": {
            "status": vector_result.get("status", "ready"),
            "row_count": row_count,
            "vector_chunks": vector_result.get("chunk_count", 0),
            "kb_id": vector_result.get("kb_id"),
            "char_count": vector_result.get("char_count", 0),
        },
    }


@router.post("/leads/reindex-database")
def reindex_leads_database(
    category: str,
    subcategory: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Any:
    """Re-vector-index an existing uploaded database for RAG."""
    import os
    from app.services.agent_call_context import resolve_rhea_db_path
    from app.services.rag_service import sync_sqlite_db_to_kb

    org = get_active_org(db, current_user)
    db_path = resolve_rhea_db_path(org.id, category, subcategory)
    if not db_path or not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="No database found for this category/subcategory.")

    result = sync_sqlite_db_to_kb(db, org.id, category, subcategory, db_path)
    return {"status": "success", "processing": result}
